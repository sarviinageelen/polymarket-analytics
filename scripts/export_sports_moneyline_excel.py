"""Export a local sports moneyline DuckDB snapshot to an auditable Excel workbook.

The workbook keeps settled performance separate from live/open/upcoming markets.
Primary Pick is inferred from the outcome with the largest cumulative BUY
notional for a wallet/game ledger. It is a compact trading-direction heuristic,
not a claim about intent.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from collections import Counter, defaultdict
from copy import copy
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import duckdb


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, "/usr/lib/python3/dist-packages")

from openpyxl import Workbook  # noqa: E402
from openpyxl.comments import Comment  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: E402
from polymarket_analytics.analytics import odds_performance  # noqa: E402


DEFAULT_EXPERIMENT_DIR = ROOT / "data/experiments/nav_wnba_2026_moneyline"
DEFAULT_OUTPUT = ROOT / "reports/generated/wnba_2026_moneyline_picks.xlsx"
PROFILE_BASE_URL = "https://polymarket.com/profile/"

NAVY = "0F172A"
BLUE = "E2E8F0"
LIGHT_BLUE = "F8FAFC"
GREEN = "DCFCE7"
GREEN_TEXT = "166534"
RED = "FEE2E2"
RED_TEXT = "991B1B"
YELLOW = "FEF3C7"
YELLOW_TEXT = "92400E"
GRAY = "F1F5F9"
GRAY_TEXT = "475569"
WHITE = "FFFFFF"
DARK = "0F172A"
THIN_GRAY = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def status_label(status: str) -> str:
    return {
        "closed": "Closed",
        "live": "Live",
        "open": "Open",
        "upcoming": "Upcoming",
        "stale_unresolved": "Stale / unresolved",
    }.get(str(status).lower(), str(status).replace("_", " ").title())


def parse_jsonish(value: Any, default: Any = None) -> Any:
    if isinstance(value, (list, dict, int, float, bool)):
        return value
    if value is None:
        return default
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return default


def finite_float(value: Any, default: float = 0.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number if math.isfinite(number) else default


def optional_float(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def parse_epoch(value: Any) -> datetime | None:
    try:
        timestamp = float(value)
    except (TypeError, ValueError):
        return None
    if timestamp <= 0:
        return None
    return datetime.fromtimestamp(timestamp, tz=timezone.utc)


def parse_date(value: Any) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def excel_datetime(value: datetime | None) -> datetime | None:
    return value.replace(tzinfo=None) if value else None


def add_profile_hyperlink(cell: Any, wallet: str) -> None:
    """Make a bettor label or wallet cell open that wallet's Polymarket profile."""

    wallet = str(wallet or "").lower()
    if not wallet:
        return
    cell.hyperlink = f"{PROFILE_BASE_URL}{wallet}"
    font = copy(cell.font)
    font.color = "0563C1"
    font.underline = "single"
    cell.font = font


def style_title(ws: Any, title: str, subtitle: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, title)
    cell.font = Font(bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    cell = ws.cell(2, 1, subtitle)
    cell.font = Font(italic=True, color="666666")
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 32


def style_header_row(ws: Any, row_number: int, start_column: int, end_column: int, fill: str = BLUE) -> None:
    for column in range(start_column, end_column + 1):
        cell = ws.cell(row_number, column)
        cell.font = Font(bold=True, color=DARK)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_number].height = 30


def fill_cell(cell: Any, fill: str, font_color: str = DARK) -> None:
    """Apply a semantic fill and an accessible text color together."""

    cell.fill = PatternFill("solid", fgColor=fill)
    font = copy(cell.font)
    font.color = font_color
    cell.font = font


def safe_table_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "", name)
    return cleaned if cleaned and cleaned[0].isalpha() else f"T{cleaned}"


def add_table(ws: Any, name: str, start_row: int, end_row: int, end_column: int) -> None:
    """Add a table; its built-in filter is the only filter definition needed."""

    # Excel does not reliably preserve a table that contains only its header
    # row. Empty ledger views should remain a normal empty worksheet instead.
    if end_row <= start_row or end_column < 1:
        return
    ref = f"A{start_row}:{get_column_letter(end_column)}{end_row}"
    table = Table(displayName=safe_table_name(name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def row_dicts(result: Any) -> list[dict[str, Any]]:
    columns = [item[0] for item in result.description]
    return [dict(zip(columns, row)) for row in result.fetchall()]


def load_candidates(experiment_dir: Path, filename: str) -> list[dict[str, Any]]:
    path = experiment_dir / "results" / filename
    with path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    for row in rows:
        row["wallet"] = str(row.get("wallet") or "").lower()
    rows.sort(
        key=lambda row: (
            finite_float(row.get("total_pnl")),
            finite_float(row.get("settled_buy_cost")),
        ),
        reverse=True,
    )
    return rows


def normalize_market(row: dict[str, Any]) -> dict[str, Any]:
    outcomes = [str(item) for item in (parse_jsonish(row.get("outcomes"), []) or [])]
    prices = [finite_float(item) for item in (parse_jsonish(row.get("outcome_prices"), []) or [])]
    if len(outcomes) != 2:
        raise ValueError(f"moneyline {row.get('condition_id')} does not have two outcomes")
    resolution = str(row.get("resolution_type") or "unresolved").lower()
    status = str(row.get("market_status") or "unknown").lower()
    if status == "unknown" and resolution in {"resolved", "tie"}:
        status = "closed"
    return {
        "condition_id": str(row.get("condition_id") or ""),
        "event_id": str(row.get("event_id") or ""),
        "event_slug": str(row.get("event_slug") or ""),
        "title": str(row.get("title") or ""),
        "event_date": parse_date(row.get("event_date")),
        "market_start": parse_epoch(row.get("market_start_ts")),
        "market_end": parse_epoch(row.get("market_end_ts")),
        "slug": str(row.get("slug") or ""),
        "outcomes": outcomes,
        "prices": prices,
        "status": status,
        "market_closed": bool(row.get("market_closed")),
        "event_closed": bool(row.get("event_closed")),
        "event_live": bool(row.get("event_live")),
        "event_ended": bool(row.get("event_ended")),
        "team_a": str(row.get("team_a") or outcomes[0]),
        "team_b": str(row.get("team_b") or outcomes[1]),
        "current_price_a": optional_float(row.get("current_price_a")),
        "current_price_b": optional_float(row.get("current_price_b")),
        "resolution": resolution,
        "winner": str(row.get("winner") or ""),
        "final_price_a": optional_float(row.get("final_price_a")),
        "final_price_b": optional_float(row.get("final_price_b")),
    }


def load_snapshot(
    experiment_dir: Path,
    wallets: set[str] | None = None,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    """Load markets and, when requested, only the ledgers needed for the workbook.

    Large sport seasons can contain hundreds of thousands of wallet/game
    ledgers. Loading the complete ledger table and then discarding almost all
    of it for the 5+/10+ candidate views needlessly multiplies memory use.
    Passing a wallet set pushes that filter into DuckDB before Python creates
    row dictionaries.
    """

    db_candidates = sorted((experiment_dir / "silver").glob("*.duckdb"))
    if not db_candidates:
        raise FileNotFoundError(f"no DuckDB found under {experiment_dir / 'silver'}")
    manifest_path = experiment_dir / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    conn = duckdb.connect(str(db_candidates[0]), read_only=True)
    try:
        markets = [normalize_market(row) for row in row_dicts(conn.execute(
            "SELECT * FROM market_dim ORDER BY event_date, market_start_ts, event_id"
        ))]
        if wallets is None:
            ledgers = row_dicts(conn.execute(
                "SELECT * FROM wallet_game_ledger ORDER BY event_date, condition_id, wallet"
            ))
        elif wallets:
            conn.execute("CREATE TEMP TABLE candidate_workbook_wallets(wallet VARCHAR)")
            conn.executemany(
                "INSERT INTO candidate_workbook_wallets VALUES (?)",
                [(wallet,) for wallet in sorted(wallets)],
            )
            ledgers = row_dicts(conn.execute(
                """
                SELECT l.*
                FROM wallet_game_ledger l
                JOIN candidate_workbook_wallets w USING (wallet)
                ORDER BY l.event_date, l.condition_id, l.wallet
                """
            ))
        else:
            ledgers = []
        pipeline_metadata = {
            row[0]: row[1]
            for row in conn.execute("SELECT key, value FROM pipeline_metadata").fetchall()
        }
    finally:
        conn.close()
    if not manifest.get("generated_at_utc"):
        manifest["generated_at_utc"] = pipeline_metadata.get("generated_at_utc", "")
    return manifest, markets, ledgers


def load_trade_stats(experiment_dir: Path, wallets: set[str]) -> dict[tuple[str, str], dict[str, Any]]:
    db_candidates = sorted((experiment_dir / "silver").glob("*.duckdb"))
    conn = duckdb.connect(str(db_candidates[0]), read_only=True)
    try:
        conn.execute("CREATE TEMP TABLE candidate_wallets(wallet VARCHAR)")
        conn.executemany("INSERT INTO candidate_wallets VALUES (?)", [(wallet,) for wallet in sorted(wallets)])
        rows = row_dicts(conn.execute(
            """
            SELECT
                t.wallet,
                t.condition_id,
                t.outcome_index,
                SUM(CASE WHEN upper(t.side) = 'BUY' THEN t.size * t.price ELSE 0 END) AS buy_notional,
                SUM(CASE WHEN upper(t.side) = 'SELL' THEN t.size * t.price ELSE 0 END) AS sell_notional
            FROM trade_fact t
            JOIN candidate_wallets c USING (wallet)
            GROUP BY t.wallet, t.condition_id, t.outcome_index
            """
        ))
    finally:
        conn.close()
    stats: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (str(row["wallet"]).lower(), str(row["condition_id"]))
        current = stats.setdefault(key, {"buy_notional": defaultdict(float), "sell_notional": defaultdict(float)})
        index = int(row.get("outcome_index") or 0)
        current["buy_notional"][index] += finite_float(row.get("buy_notional"))
        current["sell_notional"][index] += finite_float(row.get("sell_notional"))
    return stats


def pick_label(wallet: str, condition_id: str, market: dict[str, Any], trade_stats: dict[tuple[str, str], dict[str, Any]]) -> tuple[str, str]:
    stats = trade_stats.get((wallet, condition_id), {})
    buys = {index: value for index, value in stats.get("buy_notional", {}).items() if value > 1e-9}
    if not buys:
        return "Sell-only", "No BUY rows"
    if len(buys) > 1:
        return "Both / hedged", "BUY notional on both outcomes"
    index = next(iter(buys))
    return market["outcomes"][index], "Largest cumulative BUY notional"


def position_label(ledger: dict[str, Any], market: dict[str, Any]) -> str:
    shares = [finite_float(ledger.get("net_shares_a")), finite_float(ledger.get("net_shares_b"))]
    positives = [index for index, value in enumerate(shares) if value > 1e-7]
    negatives = [index for index, value in enumerate(shares) if value < -1e-7]
    if positives and not negatives:
        return f"Long {market['outcomes'][positives[0]]}" if len(positives) == 1 else "Long both"
    if negatives and not positives:
        return f"Short {market['outcomes'][negatives[0]]}" if len(negatives) == 1 else "Short both"
    if positives and negatives:
        return "Mixed / hedged"
    return "Exited / flat"


def candidate_summary(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "wallet": str(row.get("wallet") or "").lower(),
        "name": str(row.get("name") or ""),
        "pseudonym": str(row.get("pseudonym") or ""),
        "display_name": str(row.get("display_name") or row.get("name") or row.get("pseudonym") or row.get("wallet") or ""),
        "markets": int(finite_float(row.get("markets"))),
        "settled_markets": int(finite_float(row.get("settled_markets"))),
        "open_markets": int(finite_float(row.get("open_markets"))),
        "wins": int(finite_float(row.get("wins"))),
        "losses": int(finite_float(row.get("losses"))),
        "flats": int(finite_float(row.get("flats"))),
        "unsettled": int(finite_float(row.get("unsettled"))),
        "trade_count": int(finite_float(row.get("trade_count"))),
        "win_rate": finite_float(row.get("win_rate")),
        "total_pnl": finite_float(row.get("total_pnl")),
        "mark_to_market_pnl": finite_float(row.get("mark_to_market_pnl")),
        "open_exposure": finite_float(row.get("open_exposure")),
        "settled_buy_cost": finite_float(row.get("settled_buy_cost")),
        "roi": optional_float(row.get("roi")),
    }


def build_ledger_rows(
    ledgers: list[dict[str, Any]],
    markets_by_condition: dict[str, dict[str, Any]],
    summaries: dict[str, dict[str, Any]],
    trade_stats: dict[tuple[str, str], dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ledger in ledgers:
        wallet = str(ledger.get("wallet") or "").lower()
        if wallet not in summaries:
            continue
        condition_id = str(ledger.get("condition_id") or "")
        market = markets_by_condition[condition_id]
        summary = summaries[wallet]
        pick, pick_basis = pick_label(wallet, condition_id, market, trade_stats)
        settled = market["resolution"] in {"resolved", "tie"}
        if not settled:
            pick_correct = "N/A"
        elif market["resolution"] == "tie":
            pick_correct = "N/A"
        elif pick in market["outcomes"] and market["winner"]:
            pick_correct = "Yes" if pick == market["winner"] else "No"
        else:
            pick_correct = "N/A"
        rows.append({
            "Status": status_label(market["status"]),
            "Event Date": market["event_date"],
            "Team A": market["team_a"],
            "Team B": market["team_b"],
            "Matchup": f"{market['team_a']} vs {market['team_b']}",
            "Winner": market["winner"],
            "Resolution": market["resolution"].title(),
            "Current Price A": optional_float(ledger.get("price_a")),
            "Current Price B": optional_float(ledger.get("price_b")),
            "Final Price A": optional_float(ledger.get("final_price_a")),
            "Final Price B": optional_float(ledger.get("final_price_b")),
            "Bettor": summary["display_name"],
            "Wallet": wallet,
            "Settled Games": summary["settled_markets"],
            "Wins": summary["wins"],
            "Losses": summary["losses"],
            "Win %": summary["win_rate"],
            "Realized P&L": optional_float(ledger.get("realized_pnl")),
            "Mark-to-Market P&L": optional_float(ledger.get("mark_to_market_pnl")),
            "Open Exposure": optional_float(ledger.get("mark_to_market_value")) if not settled else None,
            "Primary Pick": pick,
            "Pick Correct?": pick_correct,
            "Pick Basis": pick_basis,
            "Net Position": position_label(ledger, market),
            "Ledger Result": str(ledger.get("result") or "").title(),
            "Buy Cost": finite_float(ledger.get("buy_cost")),
            "Sell Proceeds": finite_float(ledger.get("sell_proceeds")),
            "Settlement Value": optional_float(ledger.get("settlement_value")),
            "Trade Count": int(finite_float(ledger.get("trade_count"))),
            "First Trade UTC": excel_datetime(parse_epoch(ledger.get("first_trade_timestamp"))),
            "Last Trade UTC": excel_datetime(parse_epoch(ledger.get("last_trade_timestamp"))),
            "Condition ID": condition_id,
            "Event ID": str(ledger.get("event_id") or ""),
            "Event Slug": str(ledger.get("event_slug") or ""),
        })
    rows.sort(key=lambda row: (row["Event Date"] or date.max, row["Matchup"], row["Bettor"]))
    return rows


def format_cell(cell: Any, header: str, value: Any) -> None:
    cell.border = BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    if header == "Event Date" and value:
        cell.number_format = "yyyy-mm-dd"
    elif header in {"First Trade UTC", "Last Trade UTC"} and value:
        cell.number_format = "yyyy-mm-dd hh:mm"
    elif header == "Win %":
        cell.number_format = "0.00%"
    elif header in {
        "Current Price A", "Current Price B", "Final Price A", "Final Price B",
    }:
        cell.number_format = "0.000"
    elif header in {
        "Realized P&L", "Mark-to-Market P&L", "Open Exposure", "Buy Cost",
        "Sell Proceeds", "Settlement Value", "Profit", "ROI",
    }:
        cell.number_format = '$#,##0.00;[Red]-$#,##0.00'


def write_games_sheet(wb: Workbook, games: list[dict[str, Any]], season: str, scope: str) -> None:
    ws = wb.create_sheet("Games")
    headers = [
        "Game #", "Status", "Event Date", "Team A", "Team B", "Matchup",
        "Resolution", "Winner", "Current Price A", "Current Price B",
        "Final Price A", "Final Price B", "Market Closed", "Event Closed",
        "Event Live", "Event Ended", "Condition ID", "Event ID", "Event Slug",
        "Market Question", "Market End UTC",
    ]
    style_title(
        ws,
        f"{season} Full-Game Moneyline Schedule",
        f"All {len(games)} moneyline markets in the captured snapshot. {scope}",
        len(headers),
    )
    for column, header in enumerate(headers, start=1):
        ws.cell(4, column, header)
    style_header_row(ws, 4, 1, len(headers))
    for row_number, game in enumerate(games, start=5):
        values = [
            row_number - 4, status_label(game["status"]), game["event_date"], game["team_a"], game["team_b"],
            f"{game['team_a']} vs {game['team_b']}", game["resolution"].title(), game["winner"],
            game["current_price_a"], game["current_price_b"], game["final_price_a"], game["final_price_b"],
            game["market_closed"], game["event_closed"], game["event_live"], game["event_ended"],
            game["condition_id"], game["event_id"], game["event_slug"], game["title"],
            excel_datetime(game["market_end"]),
        ]
        for column, (header, value) in enumerate(zip(headers, values), start=1):
            cell = ws.cell(row_number, column, value)
            format_cell(cell, header, value)
            if header == "Status":
                fill_cell(cell, {
                    "Closed": GREEN, "Live": BLUE, "Open": YELLOW, "Upcoming": GRAY,
                    "Stale / unresolved": RED,
                }.get(str(value), LIGHT_BLUE), {
                    "Closed": GREEN_TEXT, "Live": DARK, "Open": YELLOW_TEXT,
                    "Upcoming": GRAY_TEXT, "Stale / unresolved": RED_TEXT,
                }.get(str(value), GRAY_TEXT))
            if header == "Resolution" and value == "Resolved":
                fill_cell(cell, GREEN, GREEN_TEXT)
            if header == "Resolution" and value == "Unresolved":
                fill_cell(cell, YELLOW, YELLOW_TEXT)
        ws.cell(row_number, 6).comment = Comment(
            f"Condition ID: {game['condition_id']}\n"
            f"Market status: {game['status']}\n"
            f"Resolution: {game['resolution']}\n"
            f"Current prices: {game['current_price_a']} / {game['current_price_b']}",
            "Polymarket Analytics",
        )
        ws.row_dimensions[row_number].height = 20
    add_table(ws, "GamesTable", 4, 4 + len(games), len(headers))
    ws.freeze_panes = "A5"
    widths = [9, 12, 13, 22, 22, 42, 13, 22, 14, 14, 14, 14, 14, 13, 12, 12, 68, 10, 34, 42, 21]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.sheet_view.showGridLines = False


def write_summary_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]], filter_description: str) -> None:
    ws = wb.create_sheet(name)
    headers = [
        "Bettor", "Wallet", "Pseudonym", "Markets", "Settled Games", "Open / Upcoming",
        "Wins", "Losses", "Flats", "Win %", "Realized P&L", "Mark-to-Market P&L",
        "Open Exposure", "Settled Buy Cost", "ROI", "Trade Count",
    ]
    style_title(
        ws,
        f"Bettor Summary — {filter_description}",
        "Win % uses settled non-flat games only. Realized P&L excludes open/live/upcoming markets; Open Exposure is mark-to-market value.",
        len(headers),
    )
    for column, header in enumerate(headers, start=1):
        ws.cell(4, column, header)
    style_header_row(ws, 4, 1, len(headers))
    for row_number, row in enumerate(rows, start=5):
        summary = candidate_summary(row)
        values = [
            summary["display_name"], summary["wallet"], summary["pseudonym"], summary["markets"],
            summary["settled_markets"], summary["open_markets"], summary["wins"], summary["losses"],
            summary["flats"], summary["win_rate"], summary["total_pnl"], summary["mark_to_market_pnl"],
            summary["open_exposure"], summary["settled_buy_cost"], summary["roi"], summary["trade_count"],
        ]
        for column, (header, value) in enumerate(zip(headers, values), start=1):
            cell = ws.cell(row_number, column, value)
            format_cell(cell, header, value)
            if header in {"Bettor", "Wallet"}:
                add_profile_hyperlink(cell, summary["wallet"])
            if header == "Win %" and value is not None:
                cell.number_format = "0.00%"
            if header == "ROI" and value is not None:
                cell.number_format = "0.00%"
        ws.row_dimensions[row_number].height = 20
    add_table(ws, f"{name}Table", 4, 4 + len(rows), len(headers))
    ws.freeze_panes = "A5"
    widths = [32, 45, 25, 10, 14, 15, 9, 10, 9, 10, 16, 20, 16, 17, 10, 12]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.sheet_view.showGridLines = False


def write_ledger_sheet(
    wb: Workbook,
    name: str,
    rows: list[dict[str, Any]],
    title: str,
    subtitle: str,
) -> None:
    ws = wb.create_sheet(name)
    headers = list(rows[0].keys()) if rows else [
        "Status", "Event Date", "Matchup", "Bettor", "Wallet", "Ledger Result",
    ]
    style_title(ws, title, subtitle, len(headers))
    for column, header in enumerate(headers, start=1):
        ws.cell(4, column, header)
    style_header_row(ws, 4, 1, len(headers))
    for row_number, row in enumerate(rows, start=5):
        for column, header in enumerate(headers, start=1):
            value = row.get(header)
            cell = ws.cell(row_number, column, value)
            format_cell(cell, header, value)
            if header in {"Bettor", "Wallet"}:
                add_profile_hyperlink(cell, str(row.get("Wallet") or ""))
            if header == "Pick Correct?":
                fill_cell(cell, {
                    "Yes": GREEN, "No": RED, "N/A": YELLOW,
                }.get(str(value), GRAY), {
                    "Yes": GREEN_TEXT, "No": RED_TEXT, "N/A": YELLOW_TEXT,
                }.get(str(value), GRAY_TEXT))
            if header == "Status":
                fill_cell(cell, {
                    "Closed": GREEN, "Live": BLUE, "Open": YELLOW, "Upcoming": GRAY,
                    "Stale / unresolved": RED,
                }.get(str(value), LIGHT_BLUE), {
                    "Closed": GREEN_TEXT, "Live": DARK, "Open": YELLOW_TEXT,
                    "Upcoming": GRAY_TEXT, "Stale / unresolved": RED_TEXT,
                }.get(str(value), GRAY_TEXT))
            if header == "Ledger Result" and value == "Unsettled":
                fill_cell(cell, BLUE, DARK)
        ws.row_dimensions[row_number].height = 20
    add_table(ws, f"{name}Table", 4, 4 + len(rows), len(headers))
    if rows:
        ws.freeze_panes = "A5"
    widths = {
        "Status": 12, "Event Date": 13, "Team A": 22, "Team B": 22, "Matchup": 42,
        "Winner": 22, "Resolution": 13, "Current Price A": 14, "Current Price B": 14,
        "Final Price A": 14, "Final Price B": 14, "Bettor": 32, "Wallet": 45,
        "Settled Games": 14, "Wins": 9, "Losses": 10, "Win %": 10, "Realized P&L": 16,
        "Mark-to-Market P&L": 20, "Open Exposure": 16, "Primary Pick": 23, "Pick Correct?": 14,
        "Pick Basis": 30, "Net Position": 25, "Ledger Result": 14, "Buy Cost": 16,
        "Sell Proceeds": 16, "Settlement Value": 18, "Trade Count": 12,
        "First Trade UTC": 21, "Last Trade UTC": 21, "Condition ID": 68, "Event ID": 10, "Event Slug": 34,
    }
    for column, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(column)].width = widths.get(header, 16)
    ws.sheet_view.showGridLines = False


def write_matrix_sheet(
    wb: Workbook,
    name: str,
    candidates: list[dict[str, Any]],
    ledger_rows: list[dict[str, Any]],
    games: list[dict[str, Any]],
    season: str,
    filter_description: str,
) -> None:
    ws = wb.create_sheet(name)
    static_headers = ["User", "Games", "Wins", "Losses", "Win %", "Realized P&L"]
    start_game_column = len(static_headers) + 1
    end_column = start_game_column + len(games) - 1
    style_title(
        ws,
        f"{season} Picks Matrix — {filter_description}",
        "Each game is a column. Closed games show inferred picks; open/live/upcoming cells are marked and are not scored as correct or incorrect. X in row 3 marks the last closed game; open/upcoming games begin to its right.",
        end_column,
    )
    for column, header in enumerate(static_headers, start=1):
        ws.cell(7, column, header)
        ws.merge_cells(start_row=4, start_column=column, end_row=6, end_column=column)
        ws.cell(4, column, "")
    style_header_row(ws, 7, 1, len(static_headers))
    for index, game in enumerate(games):
        column = start_game_column + index
        ws.cell(4, column, status_label(game["status"]))
        ws.cell(5, column, game["event_date"])
        ws.cell(6, column, f"{game['team_a']} vs {game['team_b']}")
        ws.cell(7, column, "Pick / State")
        for row_number in range(4, 8):
            cell = ws.cell(row_number, column)
            cell.font = Font(bold=True, color=WHITE if row_number == 4 else DARK)
            cell.fill = PatternFill("solid", fgColor=NAVY if row_number == 4 else BLUE)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(5, column).number_format = "yyyy-mm-dd"
        ws.cell(6, column).comment = Comment(
            f"Condition ID: {game['condition_id']}\n"
            f"Resolution: {game['resolution']}\n"
            f"Winner: {game['winner'] or 'none'}\n"
            f"Current prices: {game['current_price_a']} / {game['current_price_b']}",
            "Polymarket Analytics",
        )
    ledger_by_key = {(str(row["Wallet"]).lower(), str(row["Condition ID"])): row for row in ledger_rows}
    candidate_rows = [candidate_summary(row) for row in candidates]
    for row_number, candidate in enumerate(candidate_rows, start=8):
        values = [
            candidate["display_name"], candidate["settled_markets"], candidate["wins"],
            candidate["losses"], candidate["win_rate"], candidate["total_pnl"],
        ]
        for column, (header, value) in enumerate(zip(static_headers, values), start=1):
            cell = ws.cell(row_number, column, value)
            format_cell(cell, header, value)
            if column == 1:
                add_profile_hyperlink(cell, candidate["wallet"])
            if header == "Win %":
                cell.number_format = "0.00%"
        for game_index, game in enumerate(games):
            row = ledger_by_key.get((candidate["wallet"], game["condition_id"]))
            if row is None:
                continue
            cell = ws.cell(row_number, start_game_column + game_index)
            value = row["Primary Pick"]
            cell.value = value
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            if game["resolution"] == "tie" or row["Primary Pick"] == "Both / hedged":
                fill_cell(cell, YELLOW, YELLOW_TEXT)
            elif game["resolution"] not in {"resolved", "tie"}:
                fill_cell(cell, BLUE, DARK)
            elif row["Pick Correct?"] == "Yes":
                fill_cell(cell, GREEN, GREEN_TEXT)
            elif row["Pick Correct?"] == "No":
                fill_cell(cell, RED, RED_TEXT)
            else:
                fill_cell(cell, GRAY, GRAY_TEXT)
        ws.row_dimensions[row_number].height = 20
    closed_indexes = [
        index for index, game in enumerate(games)
        if game.get("status") == "closed" or game.get("resolution") in {"resolved", "tie"}
    ]
    if closed_indexes:
        marker_column = start_game_column + max(closed_indexes)
        marker = ws.cell(3, marker_column, "X")
        marker.font = Font(bold=True, size=14, color=YELLOW_TEXT)
        marker.fill = PatternFill("solid", fgColor=YELLOW)
        marker.border = BORDER
        marker.alignment = Alignment(horizontal="center", vertical="center")
        marker.comment = Comment(
            "Last closed game in this matrix. Open, live, upcoming, or unresolved games begin to the right.",
            "Polymarket Analytics",
        )
    end_row = 7 + len(candidate_rows)
    ws.auto_filter.ref = f"A7:{get_column_letter(end_column)}{end_row}"
    ws.freeze_panes = f"{get_column_letter(start_game_column)}8"
    ws.sheet_view.zoomScale = 60
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[4].height = 23
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 42
    ws.row_dimensions[7].height = 25
    widths = [24, 9, 8, 8, 9, 14]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    for column in range(start_game_column, end_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 15


def write_streaming_workbook(
    output: Path,
    manifest: dict[str, Any],
    games: list[dict[str, Any]],
    candidates_5: list[dict[str, Any]],
    candidates_10: list[dict[str, Any]],
    rows_5: list[dict[str, Any]],
    rows_10: list[dict[str, Any]],
    open_rows: list[dict[str, Any]],
    odds: dict[str, Any],
) -> None:
    """Write very wide seasons without retaining millions of cells in RAM.

    openpyxl's editable workbook model is convenient for ordinary seasons but
    keeps every styled cell in memory. NBA-sized candidate matrices can exceed
    the VPS memory ceiling even though the underlying DuckDB snapshot is
    healthy. XlsxWriter's constant-memory mode writes rows directly to the
    XLSX package while retaining the same workbook views and pick semantics.
    """

    try:
        import xlsxwriter
    except ImportError as exc:  # pragma: no cover - exercised on fresh installs
        raise RuntimeError("Large workbook export requires XlsxWriter; install the excel dependencies") from exc

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(str(output), {"constant_memory": True})
    workbook.set_properties({
        "title": f"{manifest.get('season', 'Sports')} Polymarket moneyline analysis",
        "subject": "Cached Polymarket moneyline data and bettor analysis",
        "author": "Polymarket Analytics",
    })
    fmt = {
        "title": workbook.add_format({"bold": True, "font_size": 16, "font_color": "#FFFFFF", "bg_color": "#0F172A", "border": 1, "border_color": "#0F172A"}),
        "subtitle": workbook.add_format({"italic": True, "font_color": "#475569", "text_wrap": True, "valign": "top"}),
        "label": workbook.add_format({"bold": True, "font_color": "#0F172A", "bg_color": "#E2E8F0", "border": 1, "border_color": "#CBD5E1"}),
        "header": workbook.add_format({"bold": True, "font_color": "#0F172A", "bg_color": "#E2E8F0", "border": 1, "border_color": "#CBD5E1", "text_wrap": True, "valign": "vcenter"}),
        "header_dark": workbook.add_format({"bold": True, "font_color": "#FFFFFF", "bg_color": "#0F172A", "border": 1, "border_color": "#0F172A", "text_wrap": True, "valign": "vcenter"}),
        "text": workbook.add_format({"border": 1, "border_color": "#E2E8F0", "valign": "top"}),
        "center": workbook.add_format({"border": 1, "border_color": "#E2E8F0", "align": "center", "valign": "vcenter"}),
        "date": workbook.add_format({"border": 1, "border_color": "#E2E8F0", "num_format": "yyyy-mm-dd", "align": "center"}),
        "number": workbook.add_format({"border": 1, "border_color": "#E2E8F0", "num_format": "0.00"}),
        "percent": workbook.add_format({"border": 1, "border_color": "#E2E8F0", "num_format": "0.00%"}),
        "money": workbook.add_format({"border": 1, "border_color": "#E2E8F0", "num_format": "$#,##0.00;[Red]-$#,##0.00"}),
        "link": workbook.add_format({"font_color": "#0563C1", "underline": True, "border": 1, "border_color": "#E2E8F0"}),
        "green": workbook.add_format({"font_color": "#166534", "bg_color": "#DCFCE7", "border": 1, "border_color": "#BBF7D0", "align": "center", "valign": "vcenter"}),
        "red": workbook.add_format({"font_color": "#991B1B", "bg_color": "#FEE2E2", "border": 1, "border_color": "#FECACA", "align": "center", "valign": "vcenter"}),
        "yellow": workbook.add_format({"font_color": "#92400E", "bg_color": "#FEF3C7", "border": 1, "border_color": "#FDE68A", "align": "center", "valign": "vcenter"}),
        "blue": workbook.add_format({"font_color": "#0F172A", "bg_color": "#DBEAFE", "border": 1, "border_color": "#BFDBFE", "align": "center", "valign": "vcenter"}),
        "gray": workbook.add_format({"font_color": "#475569", "bg_color": "#F1F5F9", "border": 1, "border_color": "#E2E8F0", "align": "center", "valign": "vcenter"}),
        "marker": workbook.add_format({"bold": True, "font_size": 14, "font_color": "#92400E", "bg_color": "#FEF3C7", "border": 1, "border_color": "#FDE68A", "align": "center"}),
    }

    def value_format(header: str, value: Any) -> Any:
        if value is None:
            return fmt["text"]
        if header in {"Win %", "ROI", "Actual win %", "Avg implied %", "Favorite win %", "Underdog win %", "Home win %", "Away win %"}:
            return fmt["percent"]
        if header in {"Realized P&L", "Mark-to-Market P&L", "Open Exposure", "Buy Cost", "Sell Proceeds", "Settlement Value", "Profit"}:
            return fmt["money"]
        if header in {"Event Date"}:
            return fmt["date"]
        if header in {"Current Price A", "Current Price B", "Final Price A", "Final Price B", "Favorite price", "Delta (pp)"}:
            return fmt["number"]
        return fmt["text"]

    def write_value(worksheet: Any, row: int, column: int, header: str, value: Any, cell_format: Any | None = None) -> None:
        target_format = cell_format or value_format(header, value)
        if value is None:
            worksheet.write_blank(row, column, None, target_format)
        elif isinstance(value, bool):
            worksheet.write_boolean(row, column, value, target_format)
        elif isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value)):
            worksheet.write_number(row, column, float(value), target_format)
        else:
            worksheet.write(row, column, str(value), target_format)

    def write_profile(worksheet: Any, row: int, column: int, text: Any, wallet: str) -> None:
        wallet = str(wallet or "").lower()
        if wallet:
            worksheet.write_url(row, column, f"{PROFILE_BASE_URL}{wallet}", fmt["link"], str(text or wallet))
        else:
            worksheet.write(row, column, str(text or ""), fmt["text"])

    def title(worksheet: Any, text: str, subtitle: str, end_column: int) -> None:
        worksheet.merge_range(0, 0, 0, end_column, text, fmt["title"])
        worksheet.merge_range(1, 0, 1, end_column, subtitle, fmt["subtitle"])
        worksheet.set_row(0, 26)
        worksheet.set_row(1, 34)

    season = str(manifest.get("season") or "Selected season")
    event_count = len(games)
    matrix_preview_limit = 500
    matrix_candidates_5 = candidates_5[:matrix_preview_limit]
    matrix_candidates_10 = candidates_10[:matrix_preview_limit]

    # Read Me
    worksheet = workbook.add_worksheet("Read Me")
    worksheet.set_column("A:A", 29)
    worksheet.set_column("B:B", 115)
    title(worksheet, f"{season} Polymarket Full-Game Moneyline Picks", "Streaming export for large seasons; cached API, Parquet, DuckDB, validation, and analysis remain local artifacts.", 1)
    readme_rows = [
        ("Scope", f"{len(games)} full-game moneyline markets in the configured local snapshot."),
        ("Snapshot", f"Generated/fetched at {manifest.get('generated_at_utc', 'timestamp not recorded')}; this workbook is a point-in-time local view."),
        ("Source", f"Nav1212/PolyMarketAnalytics ETL components, pinned to commit {manifest.get('source_revision', 'not recorded')}."),
        ("5+ view", f"{len(candidates_5)} bettors with at least 5 settled games, at least 70% settled non-flat win rate, and at least $1,000 settled BUY cost."),
        ("10+ view", f"{len(candidates_10)} bettors with at least 10 settled games, at least 70% settled non-flat win rate, and at least $1,000 settled BUY cost."),
        ("Matrix scope", f"The wide Picks Matrix sheets show the top {matrix_preview_limit} candidates for workbook usability. Complete candidate lists and all underlying rows remain in the local CSV, Parquet, and DuckDB artifacts."),
        ("Primary Pick", "Outcome with the largest cumulative BUY notional in that bettor/game ledger. Both outcomes = Both / hedged; no BUY rows = Sell-only."),
        ("Win %", "Settled profitable wallet × game ledgers divided by settled non-flat ledgers. Open/live/upcoming games are excluded."),
        ("Open Exposure", "Current mark-to-market value of unresolved positions using the latest cached Gamma outcome prices; it is not realized profit."),
        ("Profiles", "Bettor and Wallet cells are clickable hyperlinks to Polymarket profiles."),
        ("Odds vs results", "The Odds vs Results sheet compares the last local trade price at or before cached kickoff with resolved outcomes."),
    ]
    for index, (label, text_value) in enumerate(readme_rows, start=3):
        worksheet.write(index, 0, label, fmt["label"])
        worksheet.write(index, 1, text_value, fmt["text"])
        worksheet.set_row(index, 30 if len(text_value) > 110 else 20)
    worksheet.write_url(14, 1, "https://github.com/Nav1212/PolyMarketAnalytics", fmt["link"], "https://github.com/Nav1212/PolyMarketAnalytics")
    worksheet.write_url(15, 1, "https://docs.polymarket.com/api-reference/core/get-trades-for-a-user-or-markets", fmt["link"], "Polymarket trade API documentation")
    worksheet.hide_gridlines(2)

    # Odds vs Results
    worksheet = workbook.add_worksheet("Odds vs Results")
    title(worksheet, f"{season} Pre-match odds vs results", "The price is the latest local trade-price proxy at or before cached kickoff. It is not a sportsbook closing line.", 14)
    bands = odds.get("bands", [])
    worksheet.write(3, 0, "Favorite price bands", fmt["header_dark"])
    band_headers = ["Price band", "Games", "Favorite wins", "Favorite losses", "Actual win %", "Avg implied %", "Delta (pp)"]
    for column, header in enumerate(band_headers):
        worksheet.write(4, column, header, fmt["header"])
    for row_number, row in enumerate(bands, start=5):
        values = [row.get("band"), row.get("games"), row.get("wins"), row.get("losses"), row.get("win_rate_pct"), row.get("avg_implied_pct"), row.get("calibration_delta_pct")]
        for column, (header, value) in enumerate(zip(band_headers, values)):
            write_value(worksheet, row_number, column, header, value, fmt["green"] if header == "Delta (pp)" and value is not None and float(value) >= 0 else fmt["red"] if header == "Delta (pp)" and value is not None else None)
    team_rows = odds.get("team_rows", [])
    team_start = 7 + len(bands)
    worksheet.write(team_start, 0, "Team performance", fmt["header_dark"])
    team_headers = ["Team", "Games", "Wins", "Losses", "Win %", "Avg implied %", "Delta (pp)", "Favorite W-L", "Favorite win %", "Underdog W-L", "Underdog win %", "Home W-L", "Home win %", "Away W-L", "Away win %"]
    for column, header in enumerate(team_headers):
        worksheet.write(team_start + 1, column, header, fmt["header"])
    for row_number, row in enumerate(team_rows, start=team_start + 2):
        segments = {name: row.get(name) or {} for name in ("favorite", "underdog", "home", "away")}
        values = [row.get("team"), row.get("games"), row.get("wins"), row.get("losses"), row.get("win_rate_pct"), row.get("avg_implied_pct"), row.get("calibration_delta_pct"), _odds_segment_text(segments["favorite"]), segments["favorite"].get("win_rate_pct"), _odds_segment_text(segments["underdog"]), segments["underdog"].get("win_rate_pct"), _odds_segment_text(segments["home"]), segments["home"].get("win_rate_pct"), _odds_segment_text(segments["away"]), segments["away"].get("win_rate_pct")]
        for column, (header, value) in enumerate(zip(team_headers, values)):
            write_value(worksheet, row_number, column, header, value, fmt["green"] if header == "Delta (pp)" and value is not None and float(value) >= 0 else fmt["red"] if header == "Delta (pp)" and value is not None else None)
    worksheet.set_column(0, 0, 18)
    worksheet.set_column(1, 14, 16)
    worksheet.hide_gridlines(2)

    # Games
    worksheet = workbook.add_worksheet("Games")
    game_headers = ["Game #", "Status", "Event Date", "Team A", "Team B", "Matchup", "Resolution", "Winner", "Current Price A", "Current Price B", "Final Price A", "Final Price B", "Market Closed", "Event Closed", "Event Live", "Event Ended", "Condition ID", "Event ID", "Event Slug"]
    title(worksheet, f"{season} games", "Cached full-game moneyline market metadata and resolution state.", len(game_headers) - 1)
    for column, header in enumerate(game_headers):
        worksheet.write(3, column, header, fmt["header"])
    for row_number, game in enumerate(games, start=4):
        values = [row_number - 3, status_label(game.get("status")), game.get("event_date"), game.get("team_a"), game.get("team_b"), f"{game.get('team_a')} vs {game.get('team_b')}", str(game.get("resolution") or "").title(), game.get("winner"), game.get("current_price_a"), game.get("current_price_b"), game.get("final_price_a"), game.get("final_price_b"), game.get("market_closed"), game.get("event_closed"), game.get("event_live"), game.get("event_ended"), game.get("condition_id"), game.get("event_id"), game.get("event_slug")]
        for column, (header, value) in enumerate(zip(game_headers, values)):
            write_value(worksheet, row_number, column, header, value)
    worksheet.autofilter(3, 0, 3 + len(games), len(game_headers) - 1)
    worksheet.freeze_panes(4, 0)
    worksheet.set_column(0, 0, 9)
    worksheet.set_column(1, 1, 14)
    worksheet.set_column(2, 2, 13)
    worksheet.set_column(3, 5, 24)
    worksheet.set_column(6, 7, 16)
    worksheet.set_column(8, 11, 14)
    worksheet.set_column(16, 16, 68)
    worksheet.hide_gridlines(2)

    def stream_matrix(name: str, candidates: list[dict[str, Any]], ledger_rows: list[dict[str, Any]], description: str) -> None:
        worksheet = workbook.add_worksheet(name)
        static_headers = ["User", "Games", "Wins", "Losses", "Win %", "Realized P&L"]
        start_column = len(static_headers)
        end_column = start_column + len(games) - 1
        title(worksheet, f"{season} Picks Matrix — {description}", "Each game is a column. X in row 3 marks the last closed game; open/live/upcoming cells are kept visible but excluded from scoring.", end_column)
        for column, header in enumerate(static_headers):
            worksheet.write(6, column, header, fmt["header"])
        closed_indexes = [index for index, game in enumerate(games) if game.get("status") == "closed" or game.get("resolution") in {"resolved", "tie"}]
        for index, game in enumerate(games):
            column = start_column + index
            worksheet.write(3, column, status_label(game.get("status")), fmt["header_dark"])
            worksheet.write(4, column, game.get("event_date"), fmt["date"])
            worksheet.write(5, column, f"{game.get('team_a')} vs {game.get('team_b')}", fmt["header"])
            worksheet.write(6, column, "Pick / State", fmt["header"])
        if closed_indexes:
            worksheet.write(2, start_column + max(closed_indexes), "X", fmt["marker"])
        ledger_by_key = {(str(row["Wallet"]).lower(), str(row["Condition ID"])): row for row in ledger_rows}
        for row_number, raw_candidate in enumerate(candidates, start=7):
            candidate = candidate_summary(raw_candidate)
            write_profile(worksheet, row_number, 0, candidate["display_name"], candidate["wallet"])
            write_value(worksheet, row_number, 1, "Games", candidate["settled_markets"], fmt["center"])
            write_value(worksheet, row_number, 2, "Wins", candidate["wins"], fmt["center"])
            write_value(worksheet, row_number, 3, "Losses", candidate["losses"], fmt["center"])
            write_value(worksheet, row_number, 4, "Win %", candidate["win_rate"], fmt["percent"])
            write_value(worksheet, row_number, 5, "Realized P&L", candidate["total_pnl"], fmt["money"])
            for game_index, game in enumerate(games):
                row = ledger_by_key.get((candidate["wallet"], str(game["condition_id"])))
                if row is None:
                    worksheet.write_blank(row_number, start_column + game_index, None, fmt["gray"])
                    continue
                pick = row.get("Primary Pick")
                if game.get("resolution") == "tie" or pick == "Both / hedged":
                    cell_format = fmt["yellow"]
                elif game.get("resolution") not in {"resolved", "tie"}:
                    cell_format = fmt["blue"]
                elif row.get("Pick Correct?") == "Yes":
                    cell_format = fmt["green"]
                elif row.get("Pick Correct?") == "No":
                    cell_format = fmt["red"]
                else:
                    cell_format = fmt["gray"]
                worksheet.write(row_number, start_column + game_index, str(pick or ""), cell_format)
        end_row = 6 + len(candidates)
        worksheet.autofilter(6, 0, end_row, end_column)
        worksheet.freeze_panes(7, start_column)
        worksheet.set_column(0, 0, 24)
        worksheet.set_column(1, 5, 12)
        worksheet.set_column(start_column, end_column, 15)
        worksheet.set_row(3, 23)
        worksheet.set_row(4, 22)
        worksheet.set_row(5, 42)
        worksheet.hide_gridlines(2)

    stream_matrix("Picks Matrix (10+)", matrix_candidates_10, rows_10, f"Top {len(matrix_candidates_10)} of ≥10 settled games / ≥70% win rate")
    stream_matrix("Picks Matrix (5+)", matrix_candidates_5, rows_5, f"Top {len(matrix_candidates_5)} of ≥5 settled games / ≥70% win rate")

    def stream_ledger(name: str, rows: list[dict[str, Any]], description: str) -> None:
        worksheet = workbook.add_worksheet(name)
        headers = list(rows[0].keys()) if rows else ["Status", "Event Date", "Matchup", "Bettor", "Wallet", "Ledger Result"]
        title(worksheet, description, "One row per eligible bettor/game. Realized P&L is populated only for resolved/tie markets; open rows are unsettled.", len(headers) - 1)
        for column, header in enumerate(headers):
            worksheet.write(3, column, header, fmt["header"])
        linked_wallets: set[str] = set()
        for row_number, row in enumerate(rows, start=4):
            for column, header in enumerate(headers):
                value = row.get(header)
                wallet = str(row.get("Wallet") or "").lower()
                if header == "Bettor" and wallet and wallet not in linked_wallets:
                    write_profile(worksheet, row_number, column, value, str(row.get("Wallet") or ""))
                    linked_wallets.add(wallet)
                else:
                    write_value(worksheet, row_number, column, header, value)
        worksheet.autofilter(3, 0, 3 + len(rows), len(headers) - 1)
        worksheet.freeze_panes(4, 0)
        worksheet.set_column(0, len(headers) - 1, 16)
        worksheet.set_column(3, 3, 32)
        worksheet.set_column(4, 4, 45)
        worksheet.set_column(5, 5, 16)
        worksheet.hide_gridlines(2)

    stream_ledger("Picks Ledger (10+)", rows_10, "Picks Ledger — ≥10 settled games / ≥70% win rate")
    stream_ledger("Picks Ledger (5+)", rows_5, "Picks Ledger — ≥5 settled games / ≥70% win rate")
    stream_ledger("Open Exposure", open_rows, "Open / Live / Upcoming Exposure")

    def stream_summary(name: str, rows: list[dict[str, Any]], description: str) -> None:
        worksheet = workbook.add_worksheet(name)
        headers = ["Bettor", "Wallet", "Pseudonym", "Markets", "Settled Games", "Open / Upcoming", "Wins", "Losses", "Flats", "Win %", "Realized P&L", "Mark-to-Market P&L", "Open Exposure", "Settled Buy Cost", "ROI", "Trade Count"]
        title(worksheet, f"Bettor Summary — {description}", "Win % uses settled non-flat games only. Realized P&L excludes open/live/upcoming markets.", len(headers) - 1)
        for column, header in enumerate(headers):
            worksheet.write(3, column, header, fmt["header"])
        for row_number, raw_row in enumerate(rows, start=4):
            row = candidate_summary(raw_row)
            values = [row["display_name"], row["wallet"], row["pseudonym"], row["markets"], row["settled_markets"], row["open_markets"], row["wins"], row["losses"], row["flats"], row["win_rate"], row["total_pnl"], row["mark_to_market_pnl"], row["open_exposure"], row["settled_buy_cost"], row["roi"], row["trade_count"]]
            for column, (header, value) in enumerate(zip(headers, values)):
                if header in {"Bettor", "Wallet"}:
                    write_profile(worksheet, row_number, column, value, row["wallet"])
                else:
                    write_value(worksheet, row_number, column, header, value)
        worksheet.autofilter(3, 0, 3 + len(rows), len(headers) - 1)
        worksheet.freeze_panes(4, 0)
        worksheet.set_column(0, 0, 32)
        worksheet.set_column(1, 1, 45)
        worksheet.set_column(2, 2, 25)
        worksheet.set_column(3, 15, 15)
        worksheet.hide_gridlines(2)

    stream_summary("Bettor Summary (10+)", candidates_10, "≥10 settled games / ≥70% win rate")
    stream_summary("Bettor Summary (5+)", candidates_5, "≥5 settled games / ≥70% win rate")
    workbook.close()


def _section_heading(ws: Any, row_number: int, title: str, end_column: int) -> None:
    ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=end_column)
    cell = ws.cell(row_number, 1, title)
    fill_cell(cell, NAVY, WHITE)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_number].height = 23


def _odds_segment_text(segment: dict[str, Any] | None) -> str:
    if not segment or not segment.get("games"):
        return "—"
    return f"{segment.get('wins', 0)}-{segment.get('losses', 0)}"


def write_odds_sheet(wb: Workbook, odds: dict[str, Any], season: str) -> None:
    """Add the same pre-match price comparison used by the web dashboard."""

    ws = wb.create_sheet("Odds vs Results")
    summary = odds.get("summary", {})
    bands = odds.get("bands", [])
    team_rows = odds.get("team_rows", [])
    games = odds.get("games", [])
    end_column = 15
    style_title(
        ws,
        f"{season} Pre-match odds vs results",
        "The price is the latest local trade-price proxy at or before cached kickoff. It is not a sportsbook closing line; delta is actual win rate minus average observed price.",
        end_column,
    )
    _section_heading(ws, 4, "Favorite price bands", 7)
    band_headers = ["Price band", "Games", "Favorite wins", "Favorite losses", "Actual win %", "Avg implied %", "Delta (pp)"]
    for column, header in enumerate(band_headers, start=1):
        ws.cell(5, column, header)
    style_header_row(ws, 5, 1, len(band_headers))
    for row_number, row in enumerate(bands, start=6):
        values = [row.get("band"), row.get("games"), row.get("wins"), row.get("losses"), row.get("win_rate_pct"), row.get("avg_implied_pct"), row.get("calibration_delta_pct")]
        for column, (header, value) in enumerate(zip(band_headers, values), start=1):
            cell = ws.cell(row_number, column, value)
            format_cell(cell, header, value)
            if header in {"Actual win %", "Avg implied %"} and value is not None:
                cell.number_format = "0.00"
            if header == "Delta (pp)" and value is not None:
                fill_cell(cell, GREEN if float(value) >= 0 else RED, GREEN_TEXT if float(value) >= 0 else RED_TEXT)
        ws.row_dimensions[row_number].height = 20
    band_end = 5 + len(bands)

    team_heading = band_end + 3
    _section_heading(ws, team_heading, "Team performance", end_column)
    team_header_row = team_heading + 1
    team_headers = [
        "Team", "Games", "Wins", "Losses", "Win %", "Avg implied %", "Delta (pp)",
        "Favorite W-L", "Favorite win %", "Underdog W-L", "Underdog win %",
        "Home W-L", "Home win %", "Away W-L", "Away win %",
    ]
    for column, header in enumerate(team_headers, start=1):
        ws.cell(team_header_row, column, header)
    style_header_row(ws, team_header_row, 1, len(team_headers))
    for row_number, row in enumerate(team_rows, start=team_header_row + 1):
        segments = {name: row.get(name) for name in ("favorite", "underdog", "home", "away")}
        values = [
            row.get("team"), row.get("games"), row.get("wins"), row.get("losses"), row.get("win_rate_pct"),
            row.get("avg_implied_pct"), row.get("calibration_delta_pct"), _odds_segment_text(segments["favorite"]),
            (segments["favorite"] or {}).get("win_rate_pct"), _odds_segment_text(segments["underdog"]),
            (segments["underdog"] or {}).get("win_rate_pct"), _odds_segment_text(segments["home"]),
            (segments["home"] or {}).get("win_rate_pct"), _odds_segment_text(segments["away"]),
            (segments["away"] or {}).get("win_rate_pct"),
        ]
        for column, (header, value) in enumerate(zip(team_headers, values), start=1):
            cell = ws.cell(row_number, column, value)
            format_cell(cell, header, value)
            if header.endswith("%") and value is not None:
                cell.number_format = "0.00"
            if header == "Delta (pp)" and value is not None:
                fill_cell(cell, GREEN if float(value) >= 0 else RED, GREEN_TEXT if float(value) >= 0 else RED_TEXT)
        ws.row_dimensions[row_number].height = 20
    team_end = team_header_row + len(team_rows)
    add_table(ws, "OddsTeamTable", team_header_row, team_end, len(team_headers))

    games_heading = team_end + 3
    _section_heading(ws, games_heading, "Game-level audit", 9)
    game_header_row = games_heading + 1
    game_headers = ["Event date", "Matchup", "Favorite", "Favorite price", "Winner", "Favorite result", "Home", "Away", "Condition ID"]
    for column, header in enumerate(game_headers, start=1):
        ws.cell(game_header_row, column, header)
    style_header_row(ws, game_header_row, 1, len(game_headers))
    for row_number, game in enumerate(games, start=game_header_row + 1):
        values = [
            game.get("event_date"), game.get("title"), game.get("favorite_team"), game.get("favorite_implied_pct"),
            game.get("winner") or ("Tie" if game.get("resolution") == "tie" else None), game.get("favorite_result"),
            game.get("home_team") or "Unknown", game.get("away_team") or "Unknown", game.get("condition_id"),
        ]
        for column, (header, value) in enumerate(zip(game_headers, values), start=1):
            cell = ws.cell(row_number, column, value)
            format_cell(cell, header, value)
            if header == "Event date" and value:
                cell.number_format = "yyyy-mm-dd"
            if header == "Favorite price" and value is not None:
                cell.number_format = "0.00"
            if header == "Favorite result":
                result = str(value or "")
                fill_cell(cell, GREEN if result == "win" else RED if result == "loss" else YELLOW if result == "tie" else GRAY, GREEN_TEXT if result == "win" else RED_TEXT if result == "loss" else YELLOW_TEXT if result == "tie" else GRAY_TEXT)
        ws.row_dimensions[row_number].height = 20
    game_end = game_header_row + len(games)
    add_table(ws, "OddsGameTable", game_header_row, game_end, len(game_headers))
    ws.freeze_panes = "A6"
    widths = [18, 42, 26, 16, 23, 18, 25, 25, 68, 13, 16, 14, 16, 14, 16]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.sheet_view.showGridLines = False


def infer_events_path(experiment_dir: Path) -> Path:
    key = experiment_dir.name
    if key.startswith("nav_"):
        key = key[4:]
    if key.endswith("_moneyline"):
        key = key[:-10]
    return ROOT / "data" / "raw" / f"{key}_events.json"


def write_readme(
    wb: Workbook,
    manifest: dict[str, Any],
    games: list[dict[str, Any]],
    candidate_5: list[dict[str, Any]],
    candidate_10: list[dict[str, Any]],
) -> None:
    ws = wb.active
    ws.title = "Read Me"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 29
    ws.column_dimensions["B"].width = 115
    ws.merge_cells("A1:B1")
    ws["A1"] = f"{manifest.get('season', 'WNBA 2026')} Polymarket Full-Game Moneyline Picks"
    ws["A1"].font = Font(bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    status_counts = Counter(game["status"] for game in games)
    resolution_counts = Counter(game["resolution"] for game in games)
    min_date = min((game["event_date"] for game in games if game["event_date"]), default=None)
    max_date = max((game["event_date"] for game in games if game["event_date"]), default=None)
    generated = str(manifest.get("generated_at_utc") or manifest.get("fetched_at_utc") or "")
    season = str(manifest.get("season") or "the selected season")
    rows = [
        ("Scope", f"{len(games)} full-game moneyline markets from {min_date} through {max_date} in the local snapshot."),
        ("Snapshot", f"Generated/fetched at {generated or 'timestamp not recorded'}; {season} is represented as a point-in-time local view."),
        ("Market status", f"Closed: {status_counts['closed']}; live: {status_counts['live']}; open: {status_counts['open']}; stale/unresolved: {status_counts['stale_unresolved']}; upcoming: {status_counts['upcoming']}."),
        ("Resolution", f"Resolved: {resolution_counts['resolved']}; tie: {resolution_counts['tie']}; unresolved: {resolution_counts['unresolved']}. Unresolved markets are never treated as losses or wins."),
        ("Source", f"Nav1212/PolyMarketAnalytics ETL components, pinned to commit {manifest.get('source_revision', 'not recorded')}."),
        ("Local storage", "Raw API event metadata, trade Parquet shards, a market snapshot, DuckDB, CSV analysis outputs, and this workbook are generated locally. Raw/derived files stay out of Git; published workbooks are mirrored as GitHub Release assets."),
        ("Refresh", "Refresh metadata first, then rerun the Nav collector and DuckDB build before rerunning analysis/export. New scheduled games and status changes are picked up on refresh."),
        ("5+ view", f"{len(candidate_5)} bettors with at least 5 settled games, at least 70% settled non-flat win rate, and at least $1,000 settled BUY cost."),
        ("10+ view", f"{len(candidate_10)} bettors with at least 10 settled games, at least 70% settled non-flat win rate, and at least $1,000 settled BUY cost."),
        ("Primary Pick", "Outcome with the largest cumulative BUY notional in that bettor/game ledger. Both outcomes = Both / hedged; no BUY rows = Sell-only. This is inferred direction, not intent."),
        ("Win %", "Settled profitable wallet × game ledgers divided by settled non-flat ledgers. Open/live/upcoming games are excluded from the denominator."),
        ("Realized P&L", "BUY/SELL cash flow plus final settlement value for resolved/tie markets only. Explicit fees are not modeled."),
        ("Open Exposure", "Current mark-to-market value of unresolved positions using the latest cached Gamma outcome prices. It is not realized profit."),
        ("Profiles", "Bettor and Wallet cells are clickable hyperlinks to https://polymarket.com/profile/{wallet}."),
        ("Audit trail", "Picks Ledger sheets retain condition IDs, status, current/final prices, wallet addresses, BUY/SELL totals, position, settlement, and P&L fields."),
        ("Odds vs results", "The Odds vs Results sheet compares the last local trade price at or before cached kickoff with resolved outcomes. Home/away uses cached event ordering; delta is descriptive calibration, not a guaranteed edge."),
        ("Source repository", "https://github.com/Nav1212/PolyMarketAnalytics"),
        ("Trade API docs", "https://docs.polymarket.com/api-reference/core/get-trades-for-a-user-or-markets"),
    ]
    for row_number, (key, value) in enumerate(rows, start=3):
        ws.cell(row_number, 1, key)
        ws.cell(row_number, 2, value)
        ws.cell(row_number, 1).font = Font(bold=True, color=DARK)
        ws.cell(row_number, 1).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row_number, 1).border = BORDER
        ws.cell(row_number, 2).border = BORDER
        ws.cell(row_number, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_number].height = 35 if len(value) > 110 else 22
    for row_number in range(3, 3 + len(rows)):
        value = str(ws.cell(row_number, 2).value)
        if value.startswith("https://"):
            ws.cell(row_number, 2).hyperlink = value
            ws.cell(row_number, 2).style = "Hyperlink"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--experiment-dir", default=str(DEFAULT_EXPERIMENT_DIR))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()
    experiment_dir = Path(args.experiment_dir)
    output = Path(args.output)

    # Read the compact market dimension first. Candidate wallets are loaded
    # from CSV, then only their ledgers are fetched from DuckDB.
    manifest, games, _ = load_snapshot(experiment_dir, wallets=set())
    games_by_condition = {game["condition_id"]: game for game in games}
    candidates_5 = load_candidates(experiment_dir, "bettor_candidates_5games_70pct.csv")
    candidates_10 = load_candidates(experiment_dir, "bettor_candidates_10games_70pct.csv")
    candidate_5_summaries = {row["wallet"]: candidate_summary(row) for row in candidates_5}
    candidate_10_summaries = {row["wallet"]: candidate_summary(row) for row in candidates_10}
    union_summaries = {**candidate_5_summaries, **candidate_10_summaries}
    _, _, all_ledgers = load_snapshot(experiment_dir, wallets=set(union_summaries))
    trade_stats = load_trade_stats(experiment_dir, set(union_summaries))
    all_ledger_rows = build_ledger_rows(all_ledgers, games_by_condition, union_summaries, trade_stats)
    rows_5 = [row for row in all_ledger_rows if row["Wallet"] in candidate_5_summaries]
    rows_10 = [row for row in all_ledger_rows if row["Wallet"] in candidate_10_summaries]
    open_rows = [row for row in all_ledger_rows if row["Resolution"] not in {"Resolved", "Tie"}]
    db_candidates = sorted((experiment_dir / "silver").glob("*.duckdb"))
    odds = odds_performance(db_candidates[0], infer_events_path(experiment_dir)) if db_candidates else {"summary": {}, "bands": [], "team_rows": [], "games": []}

    output.parent.mkdir(parents=True, exist_ok=True)
    # A normal openpyxl workbook retains every cell object. Switch to the
    # constant-memory writer when a wide candidate matrix would otherwise
    # create millions of styled cells in the VPS process.
    matrix_cells = max(len(candidates_5), len(candidates_10)) * len(games)
    if matrix_cells > 2_000_000:
        write_streaming_workbook(
            output,
            manifest,
            games,
            candidates_5,
            candidates_10,
            rows_5,
            rows_10,
            open_rows,
            odds,
        )
        print(json.dumps({
            "output": str(output),
            "games": len(games),
            "candidate_5": len(candidates_5),
            "candidate_10": len(candidates_10),
            "ledgers_5": len(rows_5),
            "ledgers_10": len(rows_10),
            "open_exposure_rows": len(open_rows),
            "writer": "xlsxwriter_constant_memory",
            "matrix_cells": matrix_cells,
            "sheets": ["Read Me", "Odds vs Results", "Games", "Picks Matrix (10+)", "Picks Matrix (5+)", "Picks Ledger (10+)", "Picks Ledger (5+)", "Open Exposure", "Bettor Summary (10+)", "Bettor Summary (5+)"],
        }, indent=2))
        return 0
    workbook = Workbook()
    scope = f"Configured season window: {manifest.get('start_date')} through {manifest.get('end_date')}; future games appear as they are listed by the API."
    write_readme(workbook, manifest, games, candidates_5, candidates_10)
    write_odds_sheet(workbook, odds, str(manifest.get("season") or "Selected season"))
    write_games_sheet(workbook, games, str(manifest.get("season") or "WNBA 2026"), scope)
    season = str(manifest.get("season") or "Selected season")
    write_matrix_sheet(workbook, "Picks Matrix (10+)", candidates_10, rows_10, games, season, "≥10 settled games / ≥70% win rate")
    write_matrix_sheet(workbook, "Picks Matrix (5+)", candidates_5, rows_5, games, season, "≥5 settled games / ≥70% win rate")
    write_ledger_sheet(
        workbook,
        "Picks Ledger (10+)",
        rows_10,
        "Picks Ledger — ≥10 settled games / ≥70% win rate",
        "One row per eligible bettor/game. Realized P&L is populated only for resolved/tie markets; open rows are marked unsettled.",
    )
    write_ledger_sheet(
        workbook,
        "Picks Ledger (5+)",
        rows_5,
        "Picks Ledger — ≥5 settled games / ≥70% win rate",
        "One row per eligible bettor/game. Realized P&L is populated only for resolved/tie markets; open rows are marked unsettled.",
    )
    write_ledger_sheet(
        workbook,
        "Open Exposure",
        open_rows,
        "Open / Live / Upcoming Exposure",
        "Unresolved wallet-game ledgers only. Values are current mark-to-market exposure and P&L, not realized results.",
    )
    write_summary_sheet(workbook, "Bettor Summary (10+)", candidates_10, "≥10 settled games / ≥70% win rate")
    write_summary_sheet(workbook, "Bettor Summary (5+)", candidates_5, "≥5 settled games / ≥70% win rate")
    workbook.save(output)
    print(json.dumps({
        "output": str(output),
        "games": len(games),
        "candidate_5": len(candidates_5),
        "candidate_10": len(candidates_10),
        "ledgers_5": len(rows_5),
        "ledgers_10": len(rows_10),
        "open_exposure_rows": len(open_rows),
        "sheets": workbook.sheetnames,
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
