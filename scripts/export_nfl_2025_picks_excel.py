"""Export NFL 2025 full-game moneyline candidates and picks to Excel.

The workbook is intentionally presentation-oriented.  Its matrix sheets put
each game in a column with a four-row header (week/stage, date, matchup,
pick), matching the requested game-card layout.  The ledger sheets retain the
auditable wallet/game rows behind the matrix.

Pick rule:
    Primary Pick is the outcome with the largest cumulative BUY notional for
    a wallet/game ledger.  Buying both outcomes is labelled "Both / hedged";
    ledgers with no BUY rows are labelled "Sell-only".  This is an inferred
    trading direction, not a claim about a bettor's intent.
"""

from __future__ import annotations

import csv
import json
import sys
from collections import defaultdict
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

import pyarrow.parquet as pq


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
# The base environment supplies the workbook writer as a Debian package while
# the Nav virtualenv supplies PyArrow and the other ETL dependencies.
sys.path.insert(0, "/usr/lib/python3/dist-packages")

from openpyxl import Workbook  # noqa: E402
from openpyxl.comments import Comment  # noqa: E402
from openpyxl.formatting.rule import CellIsRule  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: E402


EXPERIMENT_DIR = ROOT / "data/experiments/nav_nfl_2025_moneyline"
EVENTS_PATH = ROOT / "data/raw/nfl_2025_events.json"
DEFAULT_OUTPUT = ROOT / "reports/generated/nfl_2025_moneyline_picks.xlsx"

NAVY = "17365D"
BLUE = "D9EAF7"
LIGHT_BLUE = "EAF3F8"
GREEN = "E2F0D9"
RED = "FCE4D6"
YELLOW = "FFF2CC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
DARK = "1F1F1F"
THIN_GRAY = Side(style="thin", color="B7B7B7")
MEDIUM_NAVY = Side(style="medium", color=NAVY)
BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def parse_jsonish(value: Any, default: Any = None) -> Any:
    if isinstance(value, (list, dict, int, float, bool)):
        return value
    if value is None:
        return default
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return default


def parse_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    text = str(value).replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def excel_datetime(value: datetime | None) -> datetime | None:
    """Excel stores datetimes without timezone metadata; values are already UTC."""

    return value.replace(tzinfo=None) if value else None


def event_date(event: dict[str, Any]) -> date:
    parsed = parse_datetime(event.get("eventDate"))
    if parsed:
        return parsed.date()
    parsed = parse_datetime(event.get("startTime"))
    if parsed:
        return parsed.date()
    raise ValueError(f"event has no usable date: {event.get('id')}")


REGULAR_WEEK_STARTS = [
    date(2025, 9, 4),
    date(2025, 9, 11),
    date(2025, 9, 18),
    date(2025, 9, 25),
    date(2025, 10, 2),
    date(2025, 10, 9),
    date(2025, 10, 16),
    date(2025, 10, 23),
    date(2025, 10, 30),
    date(2025, 11, 6),
    date(2025, 11, 13),
    date(2025, 11, 20),
    date(2025, 11, 27),
    date(2025, 12, 4),
    date(2025, 12, 11),
    date(2025, 12, 18),
    date(2025, 12, 25),
    date(2026, 1, 1),
]


def stage_label(day: date) -> tuple[str, str]:
    """Return a human-readable week/stage without relying on bad source weeks."""

    for index, start in reversed(list(enumerate(REGULAR_WEEK_STARTS, start=1))):
        if day >= start and day <= date(2026, 1, 4):
            return f"Week {index}", "Regular Season"
    if date(2026, 1, 10) <= day <= date(2026, 1, 12):
        return "Wild Card", "Playoffs"
    if date(2026, 1, 17) <= day <= date(2026, 1, 18):
        return "Divisional", "Playoffs"
    if day == date(2026, 1, 25):
        return "Conference", "Playoffs"
    if day == date(2026, 2, 8):
        return "Super Bowl", "Playoffs"
    return "Other", "NFL 2025 Series"


def load_games() -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    events = json.loads(EVENTS_PATH.read_text(encoding="utf-8"))
    games: list[dict[str, Any]] = []
    by_condition: dict[str, dict[str, Any]] = {}
    for event in events:
        day = event_date(event)
        week, stage = stage_label(day)
        for market in event.get("markets") or []:
            if market.get("sportsMarketType") != "moneyline":
                continue
            condition_id = str(market.get("conditionId") or "")
            if not condition_id:
                continue
            outcomes = [str(value) for value in (parse_jsonish(market.get("outcomes"), []) or [])]
            prices = []
            for value in parse_jsonish(market.get("outcomePrices"), []) or []:
                try:
                    prices.append(float(value))
                except (TypeError, ValueError):
                    prices.append(0.0)
            if len(outcomes) != 2:
                raise ValueError(f"moneyline {condition_id} does not have two outcomes")
            winner_index = None
            for index, price in enumerate(prices):
                if price >= 0.999:
                    winner_index = index
                    break
            winner = outcomes[winner_index] if winner_index is not None else ""
            is_tie = len(prices) == 2 and all(abs(price - 0.5) < 1e-9 for price in prices)
            resolution = "Tie" if is_tie else "Resolved" if winner else "Unresolved"
            start = parse_datetime(market.get("gameStartTime") or event.get("startTime"))
            game = {
                "game_number": 0,
                "condition_id": condition_id,
                "event_id": str(event.get("id") or ""),
                "event_slug": str(event.get("slug") or ""),
                "date": day,
                "week": week,
                "stage": stage,
                "team_a": outcomes[0],
                "team_b": outcomes[1],
                "matchup": f"{outcomes[0]} vs {outcomes[1]}",
                "winner": winner,
                "resolution": resolution,
                "outcomes": outcomes,
                "outcome_prices": prices,
                "title": str(market.get("question") or event.get("title") or ""),
                "market_close": parse_datetime(market.get("closedTime") or event.get("closedTime")),
                "kickoff": start,
            }
            games.append(game)
            by_condition[condition_id] = game
    games.sort(key=lambda row: (row["date"], row["kickoff"] or datetime.min.replace(tzinfo=timezone.utc), row["event_id"]))
    for number, game in enumerate(games, start=1):
        game["game_number"] = number
    return games, by_condition


def read_candidates(filename: str) -> tuple[list[dict[str, Any]], set[str]]:
    path = EXPERIMENT_DIR / "results" / filename
    with path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    rows.sort(key=lambda row: (float(row["total_pnl"]), float(row["buy_cost"])), reverse=True)
    wallets = {str(row["wallet"]).lower() for row in rows}
    return rows, wallets


def new_ledger(wallet: str, condition_id: str) -> dict[str, Any]:
    return {
        "wallet": wallet,
        "condition_id": condition_id,
        "name": "",
        "pseudonym": "",
        "cash_flow": 0.0,
        "buy_cost": 0.0,
        "sell_proceeds": 0.0,
        "trade_count": 0,
        "first_timestamp": None,
        "last_timestamp": None,
        "buy_notional": defaultdict(float),
        "sell_notional": defaultdict(float),
        "net_shares": defaultdict(float),
    }


def build_ledgers(wallets: set[str], games_by_condition: dict[str, dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    ledgers: dict[tuple[str, str], dict[str, Any]] = {}
    trade_dir = EXPERIMENT_DIR / "bronze" / "trades"
    for path in sorted(trade_dir.glob("*.parquet")):
        parquet = pq.ParquetFile(path)
        for batch in parquet.iter_batches(batch_size=50_000):
            for row in batch.to_pylist():
                wallet = str(row.get("proxyWallet") or "").lower()
                condition_id = str(row.get("conditionId") or "")
                if wallet not in wallets or condition_id not in games_by_condition:
                    continue
                key = (wallet, condition_id)
                ledger = ledgers.setdefault(key, new_ledger(wallet, condition_id))
                ledger["name"] = ledger["name"] or str(row.get("name") or "")
                ledger["pseudonym"] = ledger["pseudonym"] or str(row.get("pseudonym") or "")
                side = str(row.get("side") or "").upper()
                if side not in {"BUY", "SELL"}:
                    continue
                try:
                    outcome_index = int(row.get("outcomeIndex") or 0)
                    size = float(row.get("size") or 0.0)
                    price = float(row.get("price") or 0.0)
                except (TypeError, ValueError):
                    continue
                cash = size * price
                if side == "BUY":
                    ledger["cash_flow"] -= cash
                    ledger["buy_cost"] += cash
                    ledger["buy_notional"][outcome_index] += cash
                    ledger["net_shares"][outcome_index] += size
                else:
                    ledger["cash_flow"] += cash
                    ledger["sell_proceeds"] += cash
                    ledger["sell_notional"][outcome_index] += cash
                    ledger["net_shares"][outcome_index] -= size
                ledger["trade_count"] += 1
                timestamp = int(row.get("timestamp") or 0)
                if timestamp:
                    ledger["first_timestamp"] = timestamp if ledger["first_timestamp"] is None else min(ledger["first_timestamp"], timestamp)
                    ledger["last_timestamp"] = timestamp if ledger["last_timestamp"] is None else max(ledger["last_timestamp"], timestamp)
    return ledgers


def pick_label(ledger: dict[str, Any], game: dict[str, Any]) -> tuple[str, str]:
    buys = {index: value for index, value in ledger["buy_notional"].items() if value > 1e-9}
    if not buys:
        return "Sell-only", "No BUY rows"
    if len(buys) > 1:
        return "Both / hedged", "BUY notional on both outcomes"
    index = next(iter(buys))
    outcomes = game["outcomes"]
    return outcomes[index] if index < len(outcomes) else f"Outcome {index}", "Largest cumulative BUY notional"


def position_label(ledger: dict[str, Any], game: dict[str, Any]) -> str:
    positives = [index for index, value in ledger["net_shares"].items() if value > 1e-7]
    negatives = [index for index, value in ledger["net_shares"].items() if value < -1e-7]
    outcomes = game["outcomes"]
    if positives and not negatives:
        if len(positives) == 1:
            return f"Long {outcomes[positives[0]]}"
        return "Long both"
    if negatives and not positives:
        if len(negatives) == 1:
            return f"Short {outcomes[negatives[0]]}"
        return "Short both"
    if positives and negatives:
        return "Mixed / hedged"
    return "Exited / flat"


def ledger_rows(
    ledgers: dict[tuple[str, str], dict[str, Any]],
    games_by_condition: dict[str, dict[str, Any]],
    summary_by_wallet: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    rows = []
    for (wallet, condition_id), ledger in ledgers.items():
        game = games_by_condition[condition_id]
        settlement = sum(
            shares * game["outcome_prices"][index]
            for index, shares in ledger["net_shares"].items()
            if index < len(game["outcome_prices"])
        )
        pnl = ledger["cash_flow"] + settlement
        result = "Win" if pnl > 1e-9 else "Loss" if pnl < -1e-9 else "Flat"
        primary_pick, pick_basis = pick_label(ledger, game)
        winner = game["winner"]
        if game["resolution"] != "Resolved":
            pick_correct = "N/A"
        elif primary_pick in game["outcomes"] and winner:
            pick_correct = "Yes" if primary_pick == winner else "No"
        else:
            pick_correct = "N/A"
        summary = summary_by_wallet[wallet]
        rows.append(
            {
                "Week / Stage": game["week"],
                "Stage": game["stage"],
                "Date": game["date"],
                "Team A": game["team_a"],
                "Team B": game["team_b"],
                "Matchup": game["matchup"],
                "Winner": winner,
                "Resolution": game["resolution"],
                "Bettor": summary["display_name"],
                "Wallet": wallet,
                "Bettor Wins": int(summary["wins"]),
                "Bettor Losses": int(summary["losses"]),
                "Bettor Win %": float(summary["win_rate"]),
                "Bettor Profit": float(summary["total_pnl"]),
                "Primary Pick": primary_pick,
                "Pick Correct?": pick_correct,
                "Pick Basis": pick_basis,
                "Net Position": position_label(ledger, game),
                "Ledger Result": result,
                "Ledger P&L": pnl,
                "Buy Cost": ledger["buy_cost"],
                "Sell Proceeds": ledger["sell_proceeds"],
                "Settlement Value": settlement,
                "Trade Count": ledger["trade_count"],
                "First Trade UTC": ledger["first_timestamp"] or 0,
                "Last Trade UTC": ledger["last_timestamp"] or 0,
                "Condition ID": condition_id,
                "Event ID": game["event_id"],
            }
        )
    rows.sort(key=lambda row: (row["Date"], row["Matchup"], -row["Bettor Profit"], row["Bettor"]))
    return rows


def style_title(ws: Any, title: str, subtitle: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, title)
    cell.font = Font(bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    cell = ws.cell(2, 1, subtitle)
    cell.font = Font(italic=True, color="666666")
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 30


def style_header_row(ws: Any, row_number: int, start_column: int, end_column: int, fill: str = BLUE) -> None:
    for column in range(start_column, end_column + 1):
        cell = ws.cell(row_number, column)
        cell.font = Font(bold=True, color=DARK)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_table(ws: Any, name: str, start_row: int, end_row: int, end_column: int) -> None:
    if end_row < start_row:
        return
    ref = f"A{start_row}:{get_column_letter(end_column)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def write_games_sheet(wb: Workbook, games: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Games")
    headers = ["Game #", "Week / Stage", "Stage", "Date", "Team A", "Team B", "Matchup", "Winner", "Resolution", "Condition ID", "Event ID", "Event Slug", "Market Question", "Outcome Prices", "Market Close UTC", "Kickoff UTC"]
    style_title(ws, "NFL 2025 Full-Game Moneyline Schedule", "All 285 moneyline games in the captured NFL 2025 series, including postseason games.", len(headers))
    for column, header in enumerate(headers, start=1):
        ws.cell(4, column, header)
    style_header_row(ws, 4, 1, len(headers))
    for row_number, game in enumerate(games, start=5):
        values = [
            game["game_number"], game["week"], game["stage"], game["date"], game["team_a"], game["team_b"],
            game["matchup"], game["winner"], game["resolution"], game["condition_id"], game["event_id"], game["event_slug"],
            game["title"], json.dumps(game["outcome_prices"]), excel_datetime(game["market_close"]), excel_datetime(game["kickoff"]),
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row_number, column, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=column in {7, 12})
            if column == 4 and value:
                cell.number_format = "yyyy-mm-dd"
            if column in {15, 16} and value:
                cell.number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row_number, 8).fill = PatternFill("solid", fgColor=GREEN)
    add_table(ws, "GamesTable", 4, 4 + len(games), len(headers))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{4 + len(games)}"
    widths = {1: 9, 2: 15, 3: 16, 4: 13, 5: 15, 6: 15, 7: 25, 8: 15, 9: 13, 10: 68, 11: 10, 12: 30, 13: 30, 14: 18, 15: 21, 16: 21}
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.sheet_view.showGridLines = False


def write_summary_sheet(wb: Workbook, name: str, candidates: list[dict[str, Any]], filter_description: str) -> None:
    ws = wb.create_sheet(name)
    headers = ["Bettor", "Wallet", "Pseudonym", "Games", "Wins", "Losses", "Flats", "Win %", "Profit", "Buy Cost", "ROI", "Trade Count", "Active Weeks", "Positive Weeks", "Weekly Win %", "Max Weekly Drawdown"]
    style_title(ws, f"Bettor Summary — {filter_description}", "Win % is profitable settled wallet × game ledgers / non-flat ledgers; Profit is replayed gross P&L.", len(headers))
    for column, header in enumerate(headers, start=1):
        ws.cell(4, column, header)
    style_header_row(ws, 4, 1, len(headers))
    for row_number, row in enumerate(candidates, start=5):
        values = [
            row.get("name") or row.get("pseudonym") or row["wallet"], row["wallet"], row.get("pseudonym", ""),
            int(row["markets"]), int(row["wins"]), int(row["losses"]), int(row["flats"]), float(row["win_rate"]),
            float(row["total_pnl"]), float(row["buy_cost"]), float(row["roi"]), int(row["trade_count"]),
            int(row["active_weeks"]), int(row["positive_weeks"]), float(row["weekly_win_rate"]), float(row["max_weekly_drawdown"]),
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row_number, column, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=column in {1, 2, 3})
            if column in {8, 11, 15}:
                cell.number_format = "0.00%"
            if column in {9, 10, 16}:
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
    add_table(ws, f"{name.replace(' ', '').replace('(', '').replace(')', '').replace('+', 'P')}Table", 4, 4 + len(candidates), len(headers))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{4 + len(candidates)}"
    widths = [34, 45, 24, 10, 9, 10, 9, 10, 16, 16, 10, 12, 14, 15, 15, 20]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.sheet_view.showGridLines = False


def write_ledger_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]], filter_description: str) -> None:
    ws = wb.create_sheet(name)
    headers = list(rows[0].keys()) if rows else []
    style_title(ws, f"Picks Ledger — {filter_description}", "One row per eligible bettor/game moneyline ledger. Primary Pick is inferred from cumulative BUY notional.", len(headers))
    for column, header in enumerate(headers, start=1):
        ws.cell(4, column, header)
    style_header_row(ws, 4, 1, len(headers))
    for row_number, row in enumerate(rows, start=5):
        for column, header in enumerate(headers, start=1):
            value = row[header]
            cell = ws.cell(row_number, column, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=header in {"Matchup", "Primary Pick", "Net Position", "Pick Basis"})
            if header == "Date" and value:
                cell.number_format = "yyyy-mm-dd"
            elif header in {"Bettor Win %"}:
                cell.number_format = "0.00%"
            elif header in {"Bettor Profit", "Ledger P&L", "Buy Cost", "Sell Proceeds", "Settlement Value"}:
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
            elif header in {"First Trade UTC", "Last Trade UTC"} and value:
                cell.number_format = "0"
            if header == "Pick Correct?":
                if value == "Yes":
                    cell.fill = PatternFill("solid", fgColor=GREEN)
                elif value == "No":
                    cell.fill = PatternFill("solid", fgColor=RED)
                else:
                    cell.fill = PatternFill("solid", fgColor=YELLOW)
    if headers:
        add_table(ws, f"{name.replace(' ', '').replace('(', '').replace(')', '').replace('+', 'P')}Table", 4, 4 + len(rows), len(headers))
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{4 + len(rows)}"
    widths = {
        "Week / Stage": 15, "Stage": 15, "Date": 13, "Team A": 15, "Team B": 15,
        "Matchup": 25, "Winner": 15, "Resolution": 13, "Bettor": 34, "Wallet": 45, "Bettor Wins": 12,
        "Bettor Losses": 14, "Bettor Win %": 13, "Bettor Profit": 16, "Primary Pick": 18,
        "Pick Correct?": 14, "Pick Basis": 29, "Net Position": 20, "Ledger Result": 14,
        "Ledger P&L": 16, "Buy Cost": 16, "Sell Proceeds": 16, "Settlement Value": 18,
        "Trade Count": 12, "First Trade UTC": 18, "Last Trade UTC": 18, "Condition ID": 68, "Event ID": 10,
    }
    for column, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(column)].width = widths.get(header, 16)
    ws.sheet_view.showGridLines = False


def write_matrix_sheet(
    wb: Workbook,
    name: str,
    candidates: list[dict[str, Any]],
    ledgers: list[dict[str, Any]],
    games: list[dict[str, Any]],
    filter_description: str,
) -> None:
    ws = wb.create_sheet(name)
    static_headers = ["Bettor", "Wins", "Losses", "Win %", "Profit"]
    start_game_column = len(static_headers) + 1
    end_column = start_game_column + len(games) - 1
    style_title(ws, f"NFL 2025 Picks Matrix — {filter_description}", "Each game is a column. Blank means the bettor had no moneyline ledger for that game. Pick cells are inferred from cumulative BUY notional.", end_column)
    for column, header in enumerate(static_headers, start=1):
        ws.cell(7, column, header)
        ws.merge_cells(start_row=4, start_column=column, end_row=6, end_column=column)
        ws.cell(4, column, "")
    style_header_row(ws, 7, 1, len(static_headers))
    for index, game in enumerate(games):
        column = start_game_column + index
        ws.cell(4, column, game["week"])
        ws.cell(5, column, game["date"])
        ws.cell(6, column, game["matchup"])
        ws.cell(7, column, "Pick")
        for row_number in range(4, 8):
            cell = ws.cell(row_number, column)
            cell.font = Font(bold=True, color=DARK if row_number != 4 else WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY if row_number == 4 else BLUE)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(5, column).number_format = "yyyy-mm-dd"
        ws.cell(6, column).comment = Comment(
            f"Winner: {game['winner'] or 'none'}\nResolution: {game['resolution']}\nStage: {game['stage']}\nCondition ID: {game['condition_id']}",
            "Polymarket Analytics",
        )
    ledger_by_wallet_game = {(row["Wallet"], row["Matchup"], row["Date"]): row for row in ledgers}
    # Matchups can repeat in a season, so use condition IDs from the ledger rows
    # for the actual matrix mapping.
    ledger_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for row in ledgers:
        ledger_by_key[(row["Wallet"], row["Condition ID"])] = row
    candidate_rows = []
    for candidate in candidates:
        wallet = str(candidate["wallet"]).lower()
        display = candidate.get("name") or candidate.get("pseudonym") or wallet
        candidate_rows.append((wallet, display, candidate))
    for row_number, (wallet, display, candidate) in enumerate(candidate_rows, start=8):
        values = [display, int(candidate["wins"]), int(candidate["losses"]), float(candidate["win_rate"]), float(candidate["total_pnl"])]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row_number, column, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=column == 1)
            if column == 4:
                cell.number_format = "0.00%"
            if column == 5:
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
        for game_index, game in enumerate(games):
            row = ledger_by_key.get((wallet, game["condition_id"]))
            if row is None:
                continue
            column = start_game_column + game_index
            cell = ws.cell(row_number, column, row["Primary Pick"])
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if game["resolution"] == "Tie":
                cell.fill = PatternFill("solid", fgColor=YELLOW)
            elif row["Primary Pick"] == game["winner"] and row["Primary Pick"] in game["outcomes"]:
                cell.fill = PatternFill("solid", fgColor=GREEN)
            elif row["Primary Pick"] == "Both / hedged":
                cell.fill = PatternFill("solid", fgColor=YELLOW)
            elif row["Primary Pick"] in {"Sell-only", "Exited / flat"}:
                cell.fill = PatternFill("solid", fgColor=GRAY)
            elif row["Primary Pick"] in game["outcomes"]:
                cell.fill = PatternFill("solid", fgColor=RED)
            else:
                cell.fill = PatternFill("solid", fgColor=GRAY)
    end_row = 7 + len(candidate_rows)
    ws.freeze_panes = f"{get_column_letter(start_game_column)}8"
    ws.auto_filter.ref = f"A7:{get_column_letter(end_column)}{end_row}"
    ws.sheet_view.zoomScale = 65
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[4].height = 23
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 38
    ws.row_dimensions[7].height = 24
    for column, width in {1: 34, 2: 9, 3: 10, 4: 10, 5: 16}.items():
        ws.column_dimensions[get_column_letter(column)].width = width
    for column in range(start_game_column, end_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 16


def write_readme(wb: Workbook, games: list[dict[str, Any]], counts: dict[str, int]) -> None:
    ws = wb.active
    ws.title = "Read Me"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    ws.merge_cells("A1:B1")
    ws["A1"] = "NFL 2025 Polymarket Full-Game Moneyline Picks"
    ws["A1"].font = Font(bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    rows = [
        ("Scope", "285 NFL 2025 moneyline games, including the postseason games present in the captured series."),
        ("Source", "Nav1212/PolyMarketAnalytics ETL components, pinned to commit 75d70d8f1659380591c63cc28330fc3c87efde17."),
        ("Trade API", "Public trades were fetched maker-inclusive with takerOnly=false and replayed from Parquet."),
        ("Game headers", "Matrix sheets use four header rows: week/stage, date, Team A vs Team B, and Pick."),
        ("5+ view", f"{counts['candidates_5']} bettors with at least 5 games and at least 70% profitable-game rate."),
        ("10+ view", f"{counts['candidates_10']} bettors with at least 10 games and at least 70% profitable-game rate."),
        ("Primary Pick", "Outcome with the largest cumulative BUY notional in that bettor/game ledger. Both outcomes = Both / hedged; no BUY rows = Sell-only."),
        ("Win %", "Profitable settled wallet × game ledgers divided by non-flat ledgers. This is not guaranteed directional pick accuracy."),
        ("Profit", "Gross replayed P&L from BUY/SELL cash flow plus final settlement at the captured Gamma outcome prices; explicit fees are not modeled."),
        ("Tie handling", "A resolved 0.5/0.5 moneyline, such as Packers–Cowboys, is labelled Tie rather than treated as a missing winner."),
        ("Blank matrix cell", "The selected bettor had no moneyline ledger for that game."),
        ("Audit trail", "Picks Ledger sheets retain condition IDs, wallet addresses, BUY/SELL totals, final positions, settlement, and ledger P&L."),
        ("Generated", "2026-08-02 UTC"),
    ]
    for row_number, (key, value) in enumerate(rows, start=3):
        ws.cell(row_number, 1, key)
        ws.cell(row_number, 2, value)
        ws.cell(row_number, 1).font = Font(bold=True, color=DARK)
        ws.cell(row_number, 1).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row_number, 1).border = BORDER
        ws.cell(row_number, 2).border = BORDER
        ws.cell(row_number, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_number].height = 30 if len(value) > 100 else 21
    ws["B4"].hyperlink = "https://github.com/Nav1212/PolyMarketAnalytics"
    ws["B4"].style = "Hyperlink"
    ws["B5"].hyperlink = "https://docs.polymarket.com/api-reference/core/get-trades-for-a-user-or-markets"
    ws["B5"].style = "Hyperlink"


def main() -> int:
    games, games_by_condition = load_games()
    candidates_5, wallets_5 = read_candidates("bettor_candidates_5games_70pct.csv")
    candidates_10, wallets_10 = read_candidates("bettor_candidates_10games_70pct.csv")
    all_wallets = wallets_5 | wallets_10
    ledgers = build_ledgers(all_wallets, games_by_condition)

    candidate_by_wallet = {str(row["wallet"]).lower(): row for row in candidates_5 + candidates_10}
    summary_by_wallet: dict[str, dict[str, Any]] = {}
    for wallet, candidate in candidate_by_wallet.items():
        summary_by_wallet[wallet] = dict(candidate)
        summary_by_wallet[wallet]["display_name"] = candidate.get("name") or candidate.get("pseudonym") or wallet

    all_ledger_rows = ledger_rows(ledgers, games_by_condition, summary_by_wallet)
    rows_5 = [row for row in all_ledger_rows if row["Wallet"] in wallets_5]
    rows_10 = [row for row in all_ledger_rows if row["Wallet"] in wallets_10]

    output = DEFAULT_OUTPUT
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    write_readme(wb, games, {"candidates_5": len(candidates_5), "candidates_10": len(candidates_10)})
    write_games_sheet(wb, games)
    write_matrix_sheet(wb, "Picks Matrix (10+)", candidates_10, rows_10, games, "≥10 games / ≥70% profitable-game rate")
    write_matrix_sheet(wb, "Picks Matrix (5+)", candidates_5, rows_5, games, "≥5 games / ≥70% profitable-game rate")
    write_ledger_sheet(wb, "Picks Ledger (10+)", rows_10, "≥10 games / ≥70% profitable-game rate")
    write_ledger_sheet(wb, "Picks Ledger (5+)", rows_5, "≥5 games / ≥70% profitable-game rate")
    write_summary_sheet(wb, "Bettor Summary (10+)", candidates_10, "≥10 games / ≥70% profitable-game rate")
    write_summary_sheet(wb, "Bettor Summary (5+)", candidates_5, "≥5 games / ≥70% profitable-game rate")
    wb.save(output)

    print(json.dumps({
        "output": str(output),
        "games": len(games),
        "candidates_5": len(candidates_5),
        "candidates_10": len(candidates_10),
        "ledgers_5": len(rows_5),
        "ledgers_10": len(rows_10),
        "sheets": wb.sheetnames,
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
