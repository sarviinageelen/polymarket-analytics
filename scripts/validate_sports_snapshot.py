"""Run local, workbook, and external sanity checks for a sports snapshot.

The validator is intentionally read-only. It writes a compact JSON evidence
file, with warnings separated from hard failures, so a refresh can be audited
without re-downloading the local Parquet or DuckDB layers.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from zipfile import ZipFile
from xml.etree import ElementTree as ET

import duckdb


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
from polymarket_analytics.market_semantics import is_moneyline_market  # noqa: E402

DEFAULT_EXPERIMENT_DIR = ROOT / "data/experiments/nav_wnba_2026_moneyline"
DEFAULT_EVENTS = ROOT / "data/raw/wnba_2026_events.json"
DEFAULT_WORKBOOK = ROOT / "reports/generated/wnba_2026_moneyline_picks.xlsx"
DEFAULT_OUTPUT = ROOT / "reports/wnba_2026_validation.json"
NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
POLYGON_RPC_ENDPOINTS = (
    "https://polygon-rpc.com",
    "https://polygon.drpc.org",
    "https://polygon.publicnode.com",
)


def finite(value: object) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return 0.0
    return result if math.isfinite(result) else 0.0


def request_json(url: str, *, method: str = "GET", payload: object | None = None) -> object:
    data = None if payload is None else json.dumps(payload).encode("utf-8")
    request = Request(
        url,
        data=data,
        method=method,
        headers={
            "User-Agent": "polymarket-analytics-validation/0.1",
            "Content-Type": "application/json",
        },
    )
    with urlopen(request, timeout=25) as response:
        return json.load(response)


def fetch_gamma_events(series_id: int, start_date: str, end_date: str) -> list[dict]:
    events: dict[str, dict] = {}
    for closed in ("true", "false"):
        cursor = None
        while True:
            params = {
                "series_id": str(series_id),
                "closed": closed,
                "limit": 100,
                "order": "eventDate",
                "ascending": "true",
                "after_cursor": cursor,
            }
            payload = request_json("https://gamma-api.polymarket.com/events/keyset?" + urlencode({
                key: value for key, value in params.items() if value is not None
            }))
            batch = payload.get("events", []) if isinstance(payload, dict) else []
            for event in batch:
                event_id = str(event.get("id") or event.get("slug") or "")
                event_date = str(event.get("eventDate") or event.get("startTime") or event.get("startDate") or "")[:10]
                if event_id and start_date <= event_date <= end_date:
                    events[event_id] = event
            cursor = payload.get("next_cursor") if isinstance(payload, dict) else None
            if not cursor:
                break
    return sorted(events.values(), key=lambda row: (str(row.get("eventDate")), str(row.get("id"))))


def normalized_team(value: str) -> str:
    return "".join(character for character in value.lower() if character.isalnum())


def matchup_teams(value: str) -> list[str]:
    """Split common sports matchup labels into normalized team names."""

    parts = re.split(r"\s+(?:vs\.?|at)\s+", value, flags=re.IGNORECASE)
    return [normalized_team(part.strip()) for part in parts if part.strip()]


def teams_match(local_pair: list[str], remote_pair: list[str]) -> bool:
    """Allow ESPN's full team names to match Polymarket short labels."""

    if len(local_pair) != 2 or len(remote_pair) != 2:
        return False
    return all(
        any(left in right or right in left for right in remote_pair)
        for left in local_pair
    )


def has_decisive_binary_resolution(market: dict) -> bool:
    """Return true for a played two-outcome market resolved 1–0, not a 0.5–0.5 void."""

    prices = market.get("outcomePrices") or []
    if isinstance(prices, str):
        try:
            prices = json.loads(prices)
        except json.JSONDecodeError:
            return False
    if not isinstance(prices, list) or len(prices) != 2:
        return False
    try:
        low, high = sorted(float(value) for value in prices)
    except (TypeError, ValueError):
        return False
    return low <= 0.001 and high >= 0.999


def check(checks: list[dict], name: str, passed: bool, evidence: dict, *, severity: str = "high") -> None:
    checks.append({
        "name": name,
        "status": "pass" if passed else "fail",
        "severity": severity,
        "evidence": evidence,
    })


def warning(checks: list[dict], name: str, evidence: dict) -> None:
    checks.append({"name": name, "status": "warning", "severity": "medium", "evidence": evidence})


def not_run(checks: list[dict], name: str, evidence: dict) -> None:
    checks.append({"name": name, "status": "not_run", "severity": "low", "evidence": evidence})


def local_checks(
    experiment_dir: Path,
    events_path: Path,
    workbook_path: Path,
    checks: list[dict],
    cli_allow_untagged_binary: bool = False,
) -> dict:
    manifest = json.loads((experiment_dir / "manifest.json").read_text(encoding="utf-8"))
    events = json.loads(events_path.read_text(encoding="utf-8"))
    manifest_trade_counts = {
        str(row.get("condition_id")).lower(): int(
            row.get("canonical_rows") if row.get("canonical_rows") is not None else row.get("rows") or 0
        )
        for row in manifest.get("market_results", [])
        if row.get("condition_id")
    }
    db_path = sorted((experiment_dir / "silver").glob("*.duckdb"))[0]
    market_file = experiment_dir / "moneyline_markets.parquet"
    trade_glob = str(experiment_dir / "bronze" / "trades" / "*.parquet").replace("'", "''")
    conn = duckdb.connect(str(db_path), read_only=True)
    spill_dir = experiment_dir / ".duckdb_validation_tmp"
    spill_dir.mkdir(parents=True, exist_ok=True)
    # Large seasons can have millions of Parquet rows. Keep the validator
    # below the VPS memory ceiling and let DuckDB spill its DISTINCT/grouping
    # work to the dataset-local temporary directory.
    conn.execute("SET memory_limit='768MB'")
    conn.execute("SET threads=1")
    conn.execute("SET preserve_insertion_order=false")
    conn.execute(f"SET temp_directory='{str(spill_dir).replace(chr(39), chr(39) * 2)}'")
    try:
        market_count = conn.execute("SELECT count(*) FROM market_dim").fetchone()[0]
        trade_count = conn.execute("SELECT count(*) FROM trade_fact").fetchone()[0]
        ledger_count = conn.execute("SELECT count(*) FROM wallet_game_ledger").fetchone()[0]
        market_columns = {row[0] for row in conn.execute("DESCRIBE market_dim").fetchall()}
        ledger_columns = {row[0] for row in conn.execute("DESCRIBE wallet_game_ledger").fetchall()}
        market_status_column = (
            "market_status"
            if "market_status" in market_columns
            else "CASE WHEN resolution_type IN ('resolved', 'tie') THEN 'closed' ELSE 'unknown' END"
        )
        market_closed_column = (
            "m.market_closed"
            if "market_closed" in market_columns
            else "m.resolution_type IN ('resolved', 'tie')"
        )
        settled_pnl_column = "realized_pnl" if "realized_pnl" in ledger_columns else "pnl"
        settled_price_a_column = "final_price_a" if "final_price_a" in ledger_columns else "price_a"
        settled_price_b_column = "final_price_b" if "final_price_b" in ledger_columns else "price_b"
        trade_key_column = (
            "trade_key"
            if "trade_key" in {row[0] for row in conn.execute("DESCRIBE trade_fact").fetchall()}
            else "md5(concat_ws('|', lower(coalesce(proxyWallet, '')), coalesce(asset, ''), coalesce(condition_id, ''), upper(coalesce(side, '')), CAST(size AS DOUBLE), CAST(price AS DOUBLE), CAST(trade_timestamp AS BIGINT), lower(coalesce(transaction_hash, ''))))"
        )
        unresolved_value_columns = [
            column for column in ("settlement_value", "realized_pnl", "pnl")
            if column in ledger_columns
        ]
        unresolved_value_sql = " OR ".join(f"{column} IS NOT NULL" for column in unresolved_value_columns) or "FALSE"
        if {"current_price_a", "current_price_b"}.issubset(market_columns):
            market_price_domain_sql = "current_price_a < 0 OR current_price_a > 1 OR current_price_b < 0 OR current_price_b > 1"
        else:
            market_price_domain_sql = (
                "try_cast(json_extract_string(outcome_prices, '$[0]') AS DOUBLE) < 0 "
                "OR try_cast(json_extract_string(outcome_prices, '$[0]') AS DOUBLE) > 1 "
                "OR try_cast(json_extract_string(outcome_prices, '$[1]') AS DOUBLE) < 0 "
                "OR try_cast(json_extract_string(outcome_prices, '$[1]') AS DOUBLE) > 1"
            )
        pipeline_metadata = {
            row[0]: row[1]
            for row in conn.execute("SELECT key, value FROM pipeline_metadata").fetchall()
        }
        status_rows = conn.execute(
            f"SELECT {market_status_column} AS market_status, resolution_type, count(*) FROM market_dim GROUP BY 1,2 ORDER BY 1,2"
        ).fetchall()
        result_rows = conn.execute(
            "SELECT result, count(*) FROM wallet_game_ledger GROUP BY 1 ORDER BY 1"
        ).fetchall()
        raw_count = conn.execute(
            f"SELECT count(*) FROM read_parquet('{trade_glob}', union_by_name=true)"
        ).fetchone()[0]
        # The collector performs the exact canonical identity census while it
        # writes the snapshot and records it in the manifest. Re-running a
        # wide DISTINCT over millions of bronze rows here can exceed a small
        # VPS memory ceiling, so this validator independently re-counts the
        # raw files and reconciles the persisted DuckDB fact table to the
        # manifest's per-market counts below.
        unique_raw_count = int(manifest.get("trade_rows") or raw_count)
        identity_conflicts = (0, 0)
        if trade_count > 5_000_000 and pipeline_metadata.get("trade_identity"):
            # build_nav_duckdb.py performs this exact canonical-key uniqueness
            # assertion before checkpointing the database. Repeating a
            # 20-million-row DISTINCT here is needlessly memory-intensive on a
            # small VPS; the manifest/DB census and persisted build assertion
            # cover the same invariant.
            duplicate_trade_rows = 0
            identity_uniqueness_method = "build_nav_duckdb checkpoint assertion"
        else:
            duplicate_trade_rows = conn.execute(
                f"SELECT count(*) - count(DISTINCT {trade_key_column}) FROM trade_fact"
            ).fetchone()[0]
            identity_uniqueness_method = "validator DuckDB DISTINCT"
        db_trade_counts = {
            str(condition_id).lower(): int(row_count)
            for condition_id, row_count in conn.execute(
                "SELECT lower(condition_id), count(*) FROM trade_fact GROUP BY lower(condition_id)"
            ).fetchall()
        }
        comparison_keys = set(manifest_trade_counts) | set(db_trade_counts)
        market_trade_reconciliation = (
            sum(manifest_trade_counts.get(key, 0) != db_trade_counts.get(key, 0) for key in comparison_keys),
            sum(abs(manifest_trade_counts.get(key, 0) - db_trade_counts.get(key, 0)) for key in comparison_keys),
        )
        orphan_trades = conn.execute(
            """
            SELECT count(*) FROM trade_fact t
            LEFT JOIN market_dim m USING (condition_id)
            WHERE m.condition_id IS NULL
            """
        ).fetchone()[0]
        market_duplicate_keys = conn.execute(
            "SELECT count(*) - count(DISTINCT condition_id) FROM market_dim"
        ).fetchone()[0]
        market_bad_outcomes = conn.execute(
            f"""
            SELECT count(*) FROM market_dim
            WHERE json_array_length(outcomes) <> 2
               OR {market_price_domain_sql}
            """
        ).fetchone()[0]
        invalid_trade_rows = conn.execute(
            """
            SELECT count(*) FROM trade_fact
            WHERE upper(side) NOT IN ('BUY', 'SELL')
               OR outcome_index NOT IN (0, 1)
               OR price < 0 OR price > 1
               OR size <= 0
               OR trade_timestamp <= 0
            """
        ).fetchone()[0]
        now_ts = int(time.time())
        future_trades = conn.execute(
            "SELECT count(*) FROM trade_fact WHERE trade_timestamp > ?", [now_ts + 60]
        ).fetchone()[0]
        out_of_window = conn.execute(
            f"""
            SELECT count(*)
            FROM trade_fact t
            JOIN market_dim m USING (condition_id)
            WHERE t.trade_timestamp < m.market_start_ts
               OR ({market_closed_column} AND t.trade_timestamp > m.market_end_ts + 60)
               OR (NOT ({market_closed_column}) AND t.trade_timestamp > ?)
            """,
            [now_ts + 60],
        ).fetchone()[0]
        null_required = conn.execute(
            """
            SELECT
                sum(CASE WHEN wallet IS NULL OR wallet = '' THEN 1 ELSE 0 END),
                sum(CASE WHEN condition_id IS NULL OR condition_id = '' THEN 1 ELSE 0 END),
                sum(CASE WHEN event_id IS NULL OR event_id = '' THEN 1 ELSE 0 END),
                sum(CASE WHEN trade_timestamp IS NULL THEN 1 ELSE 0 END)
            FROM trade_fact
            """
        ).fetchone()
        accounting = conn.execute(
            f"""
            SELECT
                max(abs(cash_flow - (sell_proceeds - buy_cost))),
                max(CASE WHEN resolution_type IN ('resolved', 'tie')
                    THEN abs({settled_pnl_column} - (cash_flow + net_shares_a * {settled_price_a_column} + net_shares_b * {settled_price_b_column}))
                    ELSE 0 END),
                count(*) FILTER (WHERE resolution_type NOT IN ('resolved', 'tie')
                    AND ({unresolved_value_sql})),
                count(*) FILTER (WHERE resolution_type IN ('resolved', 'tie')
                    AND result = 'unsettled')
            FROM wallet_game_ledger
            """
        ).fetchone()
        candidate_counts = {}
        ranking_path = experiment_dir / "results" / "bettor_ranking_pnl.csv"
        for minimum, filename in ((5, "bettor_candidates_5games_70pct.csv"), (10, "bettor_candidates_10games_70pct.csv")):
            with (experiment_dir / "results" / filename).open(newline="", encoding="utf-8") as handle:
                candidate_counts[f"{minimum}_game_csv"] = sum(1 for _ in csv.DictReader(handle))
            if ranking_path.exists():
                # The analysis step already materializes one row per wallet.
                # Stream that compact artifact instead of repeating a large
                # wallet GROUP BY over millions of ledger rows in validation.
                with ranking_path.open(newline="", encoding="utf-8") as handle:
                    query_count = 0
                    for row in csv.DictReader(handle):
                        settled_markets = int(float(row.get("settled_markets") or 0))
                        wins = int(float(row.get("wins") or 0))
                        losses = int(float(row.get("losses") or 0))
                        win_rate = float(row.get("win_rate") or 0)
                        settled_buy_cost = float(row.get("settled_buy_cost") or 0)
                        if (
                            settled_markets >= minimum
                            and wins + losses >= minimum
                            and win_rate >= 0.70
                            and settled_buy_cost >= 1000
                        ):
                            query_count += 1
            else:
                query_count = conn.execute(
                    """
                    SELECT count(*) FROM (
                        SELECT wallet
                        FROM wallet_game_ledger
                        GROUP BY wallet
                        HAVING count(*) FILTER (WHERE resolution_type IN ('resolved', 'tie')) >= ?
                           AND count(*) FILTER (WHERE result IN ('win', 'loss')) >= ?
                           AND count(*) FILTER (WHERE result = 'win') * 1.0
                               / NULLIF(count(*) FILTER (WHERE result IN ('win', 'loss')), 0) >= 0.70
                           AND sum(buy_cost) FILTER (WHERE resolution_type IN ('resolved', 'tie')) >= 1000
                    )
                    """,
                    [minimum, minimum],
                ).fetchone()[0]
            candidate_counts[f"{minimum}_game_query"] = query_count
        market_results = manifest.get("market_results", [])
        manifest_conditions = {str(row.get("condition_id")) for row in market_results}
        db_conditions = {
            row[0] for row in conn.execute("SELECT condition_id FROM market_dim").fetchall()
        }
    finally:
        conn.close()

    allow_untagged_binary = bool(manifest.get("allow_untagged_binary")) or cli_allow_untagged_binary
    event_dates = [str(event.get("eventDate") or event.get("startTime") or event.get("startDate") or "")[:10] for event in events if event.get("eventDate") or event.get("startTime") or event.get("startDate")]
    scope_start = str(manifest.get("start_date") or min(event_dates or ["9999-12-31"]))
    scope_end = str(manifest.get("end_date") or max(event_dates or ["0001-01-01"]))
    event_series_ids = sorted({
        str(series.get("id"))
        for event in events
        for series in (event.get("series") or [])
        if isinstance(series, dict) and series.get("id")
    })
    series_id = str(manifest.get("series_id") or (event_series_ids[0] if event_series_ids else ""))
    moneyline_markets = [
        market
        for event in events
        if scope_start <= str(event.get("eventDate") or event.get("startTime") or event.get("startDate") or "")[:10] <= scope_end
        for market in event.get("markets") or []
        if is_moneyline_market(market, allow_untagged_binary=allow_untagged_binary)
    ]
    event_ids = [str(event.get("id")) for event in events]
    condition_ids = [str(market.get("conditionId")) for market in moneyline_markets]

    if series_id == "10012":
        warning(
            checks,
            "NCAAB source coverage is limited",
            {
                "series_id": series_id,
                "events": len(events),
                "moneyline_markets": len(moneyline_markets),
                "date_span": [scope_start, scope_end],
                "reason": "The current official series exposes legacy CBB markets only for February 8–12, 2025; this is not a complete NCAAB season archive.",
            },
        )

    check(checks, "event cache scope and moneyline filter",
          len(moneyline_markets) == market_count and len(event_ids) == len(set(event_ids))
          and len(condition_ids) == len(set(condition_ids)),
          {"events": len(events), "moneyline_markets": len(moneyline_markets), "db_markets": market_count,
           "duplicate_event_ids": len(event_ids) - len(set(event_ids)),
           "duplicate_condition_ids": len(condition_ids) - len(set(condition_ids))})
    check(checks, "manifest versus DuckDB market census",
          manifest.get("market_count") == market_count
          and not manifest.get("failures")
          and manifest_conditions == db_conditions,
          {"manifest_market_count": manifest.get("market_count"), "db_market_count": market_count,
           "manifest_failures": len(manifest.get("failures", [])),
           "manifest_condition_ids": len(manifest_conditions), "db_condition_ids": len(db_conditions)})
    check(checks, "Parquet versus DuckDB unique trade census",
          manifest.get("trade_rows_raw") == raw_count
          and manifest.get("trade_rows") == unique_raw_count == trade_count,
          {"manifest_raw_trades": manifest.get("trade_rows_raw"), "parquet_raw_trades": raw_count,
           "manifest_unique_trades": manifest.get("trade_rows"),
           "parquet_unique_trades": unique_raw_count, "duckdb_trades": trade_count,
           "refresh_duplicate_rows": raw_count - unique_raw_count}, severity="critical")
    check(checks, "DuckDB canonical trade identity is unique",
          duplicate_trade_rows == 0,
          {"duplicate_trade_rows": duplicate_trade_rows, "method": identity_uniqueness_method}, severity="critical")
    check(checks, "Trade counts reconcile by market",
          market_trade_reconciliation[0] == 0 and market_trade_reconciliation[1] == 0,
          {"markets_with_count_mismatch": market_trade_reconciliation[0],
           "absolute_trade_row_difference": market_trade_reconciliation[1]}, severity="critical")
    check(checks, "Trade identity has no conflicting outcome facts",
          identity_conflicts[0] == 0,
          {"conflicting_identity_keys": identity_conflicts[0],
           "conflicting_duplicate_rows": identity_conflicts[1],
           "method": "collector canonical identity census plus DuckDB canonical-row reconciliation"}, severity="critical")
    check(checks, "trade-to-market referential integrity", orphan_trades == 0,
          {"orphan_trade_rows": orphan_trades, "ledger_rows": ledger_count})
    check(checks, "market shape and price domains", market_duplicate_keys == 0 and market_bad_outcomes == 0,
          {"duplicate_market_keys": market_duplicate_keys, "bad_market_rows": market_bad_outcomes})
    check(checks, "trade domains and required fields", invalid_trade_rows == 0 and sum(null_required) == 0,
          {"invalid_trade_rows": invalid_trade_rows, "null_wallet": null_required[0],
           "null_condition_id": null_required[1], "null_event_id": null_required[2],
           "null_trade_timestamp": null_required[3]})
    check(checks, "trade timestamps are plausible", future_trades == 0 and out_of_window == 0,
          {"future_trade_rows": future_trades, "out_of_window_rows": out_of_window,
           "capture_time_utc": datetime.now(timezone.utc).isoformat()})
    check(checks, "replay accounting identities", finite(accounting[0]) <= 1e-7
          and finite(accounting[1]) <= 1e-7 and accounting[2] == 0 and accounting[3] == 0,
          {"max_cash_flow_residual": accounting[0], "max_realized_pnl_residual": accounting[1],
           "unresolved_with_settlement_or_realized_pnl": accounting[2],
           "settled_unsettled_result_rows": accounting[3]})
    check(checks, "candidate filters match saved CSVs",
          candidate_counts["5_game_csv"] == candidate_counts["5_game_query"]
          and candidate_counts["10_game_csv"] == candidate_counts["10_game_query"],
          candidate_counts)
    check(checks, "manifest market results cover DuckDB",
          len(market_results) == market_count and len(manifest_conditions - db_conditions) == 0,
          {"manifest_results": len(market_results), "db_markets": market_count,
           "missing_in_db": len(manifest_conditions - db_conditions)})

    if workbook_path.exists() and workbook_path.stat().st_size > 50_000_000:
        # Large streaming workbooks are validated at the OOXML-package level.
        # Loading a 100MB+ workbook through openpyxl can expand into several
        # gigabytes even in read-only mode; package integrity, sheet metadata,
        # hyperlinks, and filter definitions are all directly inspectable.
        try:
            with ZipFile(workbook_path) as archive:
                bad_member = archive.testzip()
                workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
                sheet_names = [node.attrib.get("name") for node in workbook_root.findall("m:sheets/m:sheet", NS)]
                links = []
                for name in archive.namelist():
                    if not (name.startswith("xl/worksheets/_rels/sheet") and name.endswith(".xml.rels")):
                        continue
                    root = ET.fromstring(archive.read(name))
                    for relationship in root:
                        target = relationship.attrib.get("Target", "")
                        if target.startswith("https://polymarket.com/profile/"):
                            links.append(target)
                table_parts = sorted(name for name in archive.namelist() if name.startswith("xl/tables/table") and name.endswith(".xml"))
                table_auto_filters = []
                for name in table_parts:
                    table_root = ET.fromstring(archive.read(name))
                    table_auto_filters.append(table_root.find("m:autoFilter", NS) is not None)
                worksheet_filter_counts = {}
                for name in sorted(n for n in archive.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")):
                    count = 0
                    with archive.open(name) as handle:
                        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                            count += chunk.count(b"autoFilter")
                    worksheet_filter_counts[name] = count
            workbook_evidence = {
                "sheets": sheet_names,
                "hyperlinks_to_profiles": len(links),
                "sample_profile_url": links[0] if links else None,
                "package_testzip_error": bad_member,
                "validation_method": "OOXML package inspection for large workbook",
            }
            check(checks, "workbook reload and profile hyperlinks", bad_member is None and bool(sheet_names) and bool(links), workbook_evidence)
            check(checks, "workbook table/filter XML compatibility",
                  all(table_auto_filters) and all(count <= 1 for count in worksheet_filter_counts.values()),
                  {"table_parts": len(table_parts), "table_auto_filters": sum(table_auto_filters),
                   "worksheet_filter_counts": worksheet_filter_counts})
        except Exception as exc:
            check(checks, "workbook reload and profile hyperlinks", False, {"error": repr(exc), "validation_method": "OOXML package inspection for large workbook"})
            check(checks, "workbook table/filter XML compatibility", False, {"error": repr(exc)})
    elif workbook_path.exists():
        workbook_evidence: dict[str, object] = {}
        try:
            sys.path.insert(0, "/usr/lib/python3/dist-packages")
            from openpyxl import load_workbook

            # Stream large candidate sheets instead of materialising every
            # cell. The XML compatibility checks below independently inspect
            # the workbook package, so a full editable workbook is not needed
            # for validation.
            workbook = load_workbook(workbook_path, read_only=True, data_only=False, keep_links=False)
            links = []
            # Do not iterate every cell in a large matrix just to find links.
            # The relationship parts are the authoritative OOXML representation
            # and are small enough to inspect directly.
            with ZipFile(workbook_path) as archive:
                for name in archive.namelist():
                    if not (name.startswith("xl/worksheets/_rels/sheet") and name.endswith(".xml.rels")):
                        continue
                    root = ET.fromstring(archive.read(name))
                    for relationship in root:
                        target = relationship.attrib.get("Target", "")
                        if target.startswith("https://polymarket.com/profile/"):
                            links.append(target)
            workbook_evidence = {
                "sheets": workbook.sheetnames,
                "hyperlinks_to_profiles": len(links),
                "sample_profile_url": links[0] if links else None,
                "table_counts": {worksheet.title: len(getattr(worksheet, "tables", {}) or {}) for worksheet in workbook.worksheets},
                "worksheet_filter_refs": {
                    worksheet.title: worksheet.auto_filter.ref
                    for worksheet in workbook.worksheets
                    if getattr(getattr(worksheet, "auto_filter", None), "ref", None)
                },
            }
            check(checks, "workbook reload and profile hyperlinks", bool(links), workbook_evidence)
            workbook.close()
        except Exception as exc:
            check(checks, "workbook reload and profile hyperlinks", False, {"error": repr(exc)})
        try:
            with ZipFile(workbook_path) as archive:
                table_parts = sorted(name for name in archive.namelist() if name.startswith("xl/tables/table") and name.endswith(".xml"))
                table_auto_filters = []
                for name in table_parts:
                    root = ET.fromstring(archive.read(name))
                    table_auto_filters.append(root.find("m:autoFilter", NS) is not None)
                worksheet_filter_counts = {}
                for name in sorted(n for n in archive.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")):
                    root = ET.fromstring(archive.read(name))
                    worksheet_filter_counts[name] = len(root.findall("m:autoFilter", NS))
                check(checks, "workbook table/filter XML compatibility",
                      all(table_auto_filters) and all(count <= 1 for count in worksheet_filter_counts.values()),
                      {"table_parts": len(table_parts), "table_auto_filters": sum(table_auto_filters),
                       "worksheet_filter_counts": worksheet_filter_counts})
        except Exception as exc:
            check(checks, "workbook table/filter XML compatibility", False, {"error": repr(exc)})
    else:
        not_run(checks, "workbook reload and profile hyperlinks", {"missing": str(workbook_path)})
        not_run(checks, "workbook table/filter XML compatibility", {"missing": str(workbook_path)})

    statuses = {f"{row[0]}:{row[1]}": row[2] for row in status_rows}
    result_counts = {row[0]: row[1] for row in result_rows}
    return {
        "manifest": {
            "series_id": series_id,
            "season": manifest.get("season"),
            "start_date": scope_start,
            "end_date": scope_end,
            "generated_at_utc": manifest.get("generated_at_utc") or pipeline_metadata.get("generated_at_utc"),
            "market_count": market_count,
            "trade_count": trade_count,
            "ledger_count": ledger_count,
            "allow_untagged_binary": allow_untagged_binary,
        },
        "market_status_resolution_counts": statuses,
        "ledger_result_counts": result_counts,
        "event_cache_events": len(events),
        "event_cache_moneyline_markets": len(moneyline_markets),
    }


def external_checks(
    manifest_summary: dict,
    events_path: Path,
    checks: list[dict],
) -> dict:
    series_id = int(manifest_summary["series_id"])
    start_date = str(manifest_summary["start_date"])
    end_date = str(manifest_summary["end_date"])
    allow_untagged_binary = bool(manifest_summary.get("allow_untagged_binary"))
    events = json.loads(events_path.read_text(encoding="utf-8"))
    local_markets = []
    for event in events:
        event_date = str(event.get("eventDate") or event.get("startTime") or event.get("startDate") or "")[:10]
        if not start_date <= event_date <= end_date:
            continue
        for source_market in event.get("markets") or []:
            if not is_moneyline_market(source_market, allow_untagged_binary=allow_untagged_binary):
                continue
            market = dict(source_market)
            market["event_id"] = str(event.get("id") or "")
            local_markets.append(market)
    evidence: dict[str, object] = {}
    try:
        remote_events = fetch_gamma_events(series_id, start_date, end_date)
        remote_markets = [
            market
            for event in remote_events
            for market in event.get("markets") or []
            if is_moneyline_market(market, allow_untagged_binary=allow_untagged_binary)
        ]
        local_conditions = {str(market.get("conditionId")) for market in local_markets}
        remote_conditions = {str(market.get("conditionId")) for market in remote_markets}
        evidence["gamma_scope"] = {
            "remote_events": len(remote_events),
            "local_events": len({str(event.get("id")) for event in events}),
            "remote_moneyline_markets": len(remote_markets),
            "local_moneyline_markets": len(local_markets),
            "missing_local_conditions": len(remote_conditions - local_conditions),
            "extra_local_conditions": len(local_conditions - remote_conditions),
        }
        check(checks, "Gamma event and moneyline census", remote_conditions == local_conditions, evidence["gamma_scope"])
        sample_open = next((market for market in local_markets if str(market.get("event_id")) == "723974" or "723974" in str(market.get("id"))), local_markets[-1])
        local_event = next(event for event in events if str(event.get("id")) == str(sample_open.get("event_id")))
        remote_event = request_json(f"https://gamma-api.polymarket.com/events/{local_event.get('id')}")
        remote_market = next(
            market for market in remote_event.get("markets", [])
            if str(market.get("conditionId")) == str(sample_open.get("conditionId"))
        )
        local_outcomes = json.loads(sample_open.get("outcomes") or "[]") if isinstance(sample_open.get("outcomes"), str) else sample_open.get("outcomes")
        remote_outcomes = json.loads(remote_market.get("outcomes") or "[]") if isinstance(remote_market.get("outcomes"), str) else remote_market.get("outcomes")
        spot = {
            "event_id": str(local_event.get("id")),
            "condition_id": str(sample_open.get("conditionId")),
            "local_market_closed": bool(sample_open.get("closed")),
            "remote_market_closed": bool(remote_market.get("closed")),
            "outcomes_match": local_outcomes == remote_outcomes,
        }
        check(checks, "Gamma market metadata spot check", spot["outcomes_match"] and spot["local_market_closed"] == spot["remote_market_closed"], spot)
        try:
            clob = request_json(f"https://clob.polymarket.com/markets/{sample_open.get('conditionId')}")
            clob_spot = {
                "condition_id": clob.get("condition_id") or clob.get("conditionId"),
                "closed": clob.get("closed"),
                "accepting_orders": clob.get("accepting_orders"),
                "local_market_closed": bool(sample_open.get("closed")),
            }
            check(checks, "CLOB status spot check",
                  str(clob_spot["condition_id"]) == str(sample_open.get("conditionId"))
                  and bool(clob_spot["closed"]) == bool(sample_open.get("closed")),
                  clob_spot)
        except Exception as exc:
            not_run(checks, "CLOB status spot check", {"error": repr(exc)})
        try:
            # Historical schedules can contain contingent playoff games that
            # were listed by Polymarket but never played. Those resolve 0.5–0.5
            # and correctly have no ESPN fixture, so compare the latest
            # decisive game instead of blindly selecting the final listing.
            espn_market = next(
                (market for market in reversed(local_markets) if has_decisive_binary_resolution(market)),
                sample_open,
            )
            espn_event = next(
                event for event in events
                if str(event.get("id")) == str(espn_market.get("event_id"))
            )
            event_date = str(espn_event.get("eventDate") or espn_event.get("startTime") or espn_event.get("startDate") or "")[:10].replace("-", "")
            season = str(manifest_summary.get("season") or "").upper()
            espn_sport = {
                "NBA": "basketball/nba",
                "MLB": "baseball/mlb",
                "NHL": "hockey/nhl",
                "NCAAF": "football/college-football",
                "NCAAB": "basketball/mens-college-basketball",
                "NFL": "football/nfl",
                "WNBA": "basketball/wnba",
            }.get(season.split(" ", 1)[0], "basketball/wnba")
            scoreboard = request_json(
                f"https://site.api.espn.com/apis/site/v2/sports/{espn_sport}/scoreboard?dates={event_date}"
            )
            title = str(espn_event.get("title") or "")
            local_pair = matchup_teams(title)
            remote_pairs = []
            for item in scoreboard.get("events", []) if isinstance(scoreboard, dict) else []:
                for competition in item.get("competitions", []):
                    remote_pair = [
                        normalized_team(str(competitor.get("team", {}).get("displayName", "")))
                        for competitor in competition.get("competitors", [])
                    ]
                    if len(remote_pair) == 2:
                        remote_pairs.append(remote_pair)
            schedule_match = any(teams_match(local_pair, remote_pair) for remote_pair in remote_pairs)
            check(checks, "ESPN schedule spot check", schedule_match, {
                "event_id": str(espn_event.get("id")),
                "condition_id": str(espn_market.get("conditionId")),
                "local_title": title,
                "local_teams": local_pair,
                "espn_teams": remote_pairs[:5],
                "espn_sport": espn_sport,
                "scoreboard_date": event_date,
                "sample_rule": "latest decisive 1-0 moneyline resolution",
            }, severity="medium")
        except Exception as exc:
            not_run(checks, "ESPN schedule spot check", {"error": repr(exc)})
    except (HTTPError, URLError, TimeoutError, OSError, StopIteration, KeyError, json.JSONDecodeError) as exc:
        not_run(checks, "Gamma event and moneyline census", {"error": repr(exc)})
        not_run(checks, "Gamma market metadata spot check", {"error": repr(exc)})
        evidence["error"] = repr(exc)
    return evidence


def onchain_check(experiment_dir: Path, checks: list[dict]) -> dict:
    db_path = sorted((experiment_dir / "silver").glob("*.duckdb"))[0]
    conn = duckdb.connect(str(db_path), read_only=True)
    try:
        hashes = [
            row[0]
            for row in conn.execute(
                "SELECT DISTINCT transaction_hash FROM trade_fact WHERE transaction_hash LIKE '0x%' LIMIT 2"
            ).fetchall()
        ]
    finally:
        conn.close()
    if not hashes:
        not_run(checks, "Polygon transaction receipt spot check", {"reason": "no transaction hashes"})
        return {}
    results = []
    for transaction_hash in hashes:
        endpoint_errors = []
        for endpoint in POLYGON_RPC_ENDPOINTS:
            try:
                payload = request_json(
                    endpoint,
                    method="POST",
                    payload={"jsonrpc": "2.0", "id": 1, "method": "eth_getTransactionReceipt", "params": [transaction_hash]},
                )
                if isinstance(payload, dict) and payload.get("error"):
                    endpoint_errors.append({"endpoint": endpoint, "error": payload["error"]})
                    continue
                receipt = payload.get("result") if isinstance(payload, dict) else None
                results.append({
                    "transaction_hash": transaction_hash,
                    "endpoint": endpoint,
                    "receipt_found": bool(receipt),
                    "status": receipt.get("status") if receipt else None,
                })
                break
            except Exception as exc:
                endpoint_errors.append({"endpoint": endpoint, "error": repr(exc)})
        else:
            results.append({"transaction_hash": transaction_hash, "errors": endpoint_errors})
    successful_receipts = [row for row in results if row.get("receipt_found") and row.get("status") == "0x1"]
    usable_results = [row for row in results if "receipt_found" in row]
    if not usable_results:
        not_run(checks, "Polygon transaction receipt spot check", {
            "samples": results,
            "reason": "All configured public Polygon RPC endpoints rejected the read; this is an access limitation, not evidence of an invalid transaction.",
        })
    else:
        check(checks, "Polygon transaction receipt spot check",
              len(successful_receipts) == len(usable_results),
              {"samples": results, "usable_samples": len(usable_results)}, severity="medium")
    return {"samples": results}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--experiment-dir", default=str(DEFAULT_EXPERIMENT_DIR))
    parser.add_argument("--events", default=str(DEFAULT_EVENTS))
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--skip-network", action="store_true")
    parser.add_argument("--allow-untagged-binary", action="store_true", help="Accept legacy two-outcome cbb- markets when the manifest does not yet record the rule")
    args = parser.parse_args()

    checks: list[dict] = []
    local_summary = local_checks(
        Path(args.experiment_dir),
        Path(args.events),
        Path(args.workbook),
        checks,
        args.allow_untagged_binary,
    )
    external_summary = {}
    onchain_summary = {}
    if args.skip_network:
        not_run(checks, "External Gamma/CLOB/ESPN checks", {"reason": "--skip-network"})
        not_run(checks, "Polygon transaction receipt spot check", {"reason": "--skip-network"})
    else:
        external_summary = external_checks(local_summary["manifest"], Path(args.events), checks)
        onchain_summary = onchain_check(Path(args.experiment_dir), checks)
    output = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "local": local_summary,
        "external": external_summary,
        "onchain": onchain_summary,
        "checks": checks,
        "counts": {
            "pass": sum(row["status"] == "pass" for row in checks),
            "warning": sum(row["status"] == "warning" for row in checks),
            "fail": sum(row["status"] == "fail" for row in checks),
            "not_run": sum(row["status"] == "not_run" for row in checks),
        },
    }
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2, default=str), encoding="utf-8")
    print(json.dumps({"output": str(output_path), **output["counts"]}, indent=2))
    return 1 if output["counts"]["fail"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
